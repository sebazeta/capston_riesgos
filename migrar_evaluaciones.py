"""
Migración de estructura EVALUACIONES
Ajusta columnas al nuevo formato requerido por app_final.py
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

EXCEL_PATH = "matriz_riesgos_v2.xlsx"

# Nuevas columnas requeridas
NUEVAS_COLUMNAS = [
    "ID_Evaluacion",
    "Nombre", 
    "Descripcion",
    "Fecha_Creacion",
    "Responsable",
    "Estado",
    "Origen_Re_Evaluacion"
]

print("🔄 Migrando estructura de EVALUACIONES...")

# Leer datos actuales
wb = load_workbook(EXCEL_PATH)
df_actual = pd.read_excel(EXCEL_PATH, sheet_name="EVALUACIONES")

print(f"📊 Registros encontrados: {len(df_actual)}")

# Crear DataFrame con nueva estructura
df_nuevo = pd.DataFrame(columns=NUEVAS_COLUMNAS)

for _, row in df_actual.iterrows():
    nuevo_registro = {
        "ID_Evaluacion": row.get("ID_Evaluacion", ""),
        "Nombre": row.get("Nombre", ""),
        "Descripcion": row.get("Descripcion", ""),
        "Fecha_Creacion": row.get("Fecha", ""),  # Mapear Fecha → Fecha_Creacion
        "Responsable": "",  # Nueva columna (vacía por ahora)
        "Estado": "En Progreso" if row.get("Estado", "") == "Activa" else row.get("Estado", "En Progreso"),
        "Origen_Re_Evaluacion": ""  # Nueva columna
    }
    df_nuevo = pd.concat([df_nuevo, pd.DataFrame([nuevo_registro])], ignore_index=True)

# Eliminar hoja vieja y crear nueva
print("🗑️ Eliminando hoja antigua...")
if "EVALUACIONES" in wb.sheetnames:
    del wb["EVALUACIONES"]

# Crear nueva hoja
print("➕ Creando hoja con nueva estructura...")
ws = wb.create_sheet("EVALUACIONES", 1)  # Insertar después de PORTADA

# Escribir headers
for col_num, column_title in enumerate(NUEVAS_COLUMNAS, 1):
    ws.cell(row=1, column=col_num, value=column_title)

# Escribir datos
for r_idx, row in enumerate(dataframe_to_rows(df_nuevo, index=False, header=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Guardar
wb.save(EXCEL_PATH)
wb.close()

print(f"✅ Migración completada")
print(f"📊 Columnas nuevas: {', '.join(NUEVAS_COLUMNAS)}")
print(f"📈 {len(df_nuevo)} registros migrados")

# Verificar
df_verificar = pd.read_excel(EXCEL_PATH, sheet_name="EVALUACIONES")
print(f"\n🔍 Verificación:")
print(f"  Columnas: {list(df_verificar.columns)}")
print(f"  Registros: {len(df_verificar)}")
