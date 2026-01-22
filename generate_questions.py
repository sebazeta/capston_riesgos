import json
import re
import datetime as dt
from typing import List, Dict, Any

import pandas as pd
import requests
from openpyxl import load_workbook

EXCEL_PATH = "matriz_riesgos_v2.xlsx"
OLLAMA_URL = "http://localhost:11434/api/generate"

# Cambia aquí el modelo si quieres: llama3 / mistral / phi3
DEFAULT_MODEL = "llama3"

def read_sheet(sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(EXCEL_PATH, sheet_name=sheet_name)

def append_rows(sheet_name: str, rows: List[Dict[str, Any]]):
    wb = load_workbook(EXCEL_PATH)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]

    for r in rows:
        ws.append([r.get(h) for h in headers])

    wb.save(EXCEL_PATH)

def upsert_eval(sheet_name: str, id_cols: List[str], row: Dict[str, Any]):
    """
    Actualiza si existe una fila que coincide en todas las columnas id_cols,
    caso contrario inserta.
    """
    wb = load_workbook(EXCEL_PATH)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    col_idx = {h: i+1 for i, h in enumerate(headers)}

    # buscar fila existente
    target_row = None
    for r in range(2, ws.max_row + 1):
        match = True
        for k in id_cols:
            if ws.cell(row=r, column=col_idx[k]).value != row.get(k):
                match = False
                break
        if match:
            target_row = r
            break

    if target_row is None:
        ws.append([row.get(h) for h in headers])
    else:
        for h in headers:
            ws.cell(row=target_row, column=col_idx[h]).value = row.get(h)

    wb.save(EXCEL_PATH)

def ollama_generate(model: str, prompt: str) -> str:
    r = requests.post(
        OLLAMA_URL,
        json={"model": model, "prompt": prompt, "stream": False},
        timeout=90
    )
    r.raise_for_status()
    return r.json().get("response", "")

def extract_json(text: str):
    """
    Ollama a veces devuelve texto extra. Intentamos extraer el bloque JSON.
    """
    # intenta capturar el primer array JSON [...]
    m = re.search(r"\[\s*{.*}\s*\]", text, re.DOTALL)
    if not m:
        return None
    try:
        return json.loads(m.group(0))
    except Exception:
        return None

def base_questions_for_asset(bank: pd.DataFrame, tipo_activo: str, max_n=12) -> pd.DataFrame:
    """
    Selección base: preguntas del banco que coinciden por tipo,
    si no hay suficientes, toma las generales (sin tipo o cualquier).
    """
    b = bank.copy()

    # normalizar
    b["Tipo_Activo"] = b["Tipo_Activo"].fillna("").astype(str)

    match = b[b["Tipo_Activo"].str.lower() == str(tipo_activo).lower()]
    if len(match) >= max_n:
        return match.head(max_n)

    # si no hay suficientes, completa con otras
    rest = b[~b.index.isin(match.index)].head(max_n - len(match))
    return pd.concat([match, rest], ignore_index=True)

def build_prompt(asset: Dict[str, Any], base_qs: pd.DataFrame, n_final: int) -> str:
    base_list = []
    for _, row in base_qs.iterrows():
        base_list.append({
            "ID_Pregunta": str(row["ID_Pregunta"]),
            "Dimension": str(row["Dimension(D/I/C)"]),
            "Pregunta": str(row["Pregunta"]),
            "Tipo_Respuesta": str(row["Tipo_Respuesta(0/1|1-5)"]),
            "Peso": int(row["Peso(1-5)"]),
        })

    return f"""
Eres un analista de riesgos MAGERIT e ISO 27002:2022. 
Tu objetivo: generar un cuestionario específico para el activo, pero de forma CONTROLADA y AUDITABLE.

REGLAS:
1) Debes devolver ÚNICAMENTE un JSON válido: una lista de objetos.
2) Cada objeto debe tener estas claves EXACTAS:
   - ID_Pregunta (string)  -> si viene del banco, reutiliza su ID. Si es nueva: usa "IA-###".
   - Pregunta (string)
   - Tipo_Respuesta (string) -> SOLO "0/1" o "1-5"
   - Peso (integer 1-5)
   - Dimension (string) -> SOLO "D" o "I" o "C"
   - Fuente (string) -> "Base" o "IA"
3) Máximo {n_final} preguntas. No más.
4) Usa PRIORIDAD:
   - Primero selecciona y adapta del banco.
   - Solo agrega preguntas nuevas si el activo lo requiere (máximo 4 nuevas).
5) Nada de texto adicional fuera del JSON.

CONTEXTO DEL ACTIVO:
{json.dumps(asset, ensure_ascii=False)}

BANCO DE PREGUNTAS DISPONIBLE (resumen):
{json.dumps(base_list, ensure_ascii=False)}

DEVUELVE SOLO EL JSON.
""".strip()

def validate_questions(qs: List[Dict[str, Any]], n_final: int) -> List[Dict[str, Any]]:
    if not isinstance(qs, list):
        return []
    cleaned = []
    for item in qs:
        if not isinstance(item, dict):
            continue
        required = ["ID_Pregunta", "Pregunta", "Tipo_Respuesta", "Peso", "Dimension", "Fuente"]
        if any(k not in item for k in required):
            continue

        tr = str(item["Tipo_Respuesta"]).strip()
        dim = str(item["Dimension"]).strip().upper()
        fuente = str(item["Fuente"]).strip()

        if tr not in ["0/1", "1-5"]:
            continue
        if dim not in ["D", "I", "C"]:
            continue

        try:
            peso = int(item["Peso"])
        except:
            continue
        if peso < 1 or peso > 5:
            continue

        cleaned.append({
            "ID_Pregunta": str(item["ID_Pregunta"]).strip(),
            "Pregunta": str(item["Pregunta"]).strip(),
            "Tipo_Respuesta": tr,
            "Peso": peso,
            "Dimension": dim,
            "Fuente": "IA" if fuente.lower().startswith("ia") else "Base"
        })

    return cleaned[:n_final]

def main():
    # 1) leer activos
    inv = read_sheet("INVENTARIO_ACTIVOS")
    if inv.empty:
        print("No hay activos en INVENTARIO_ACTIVOS.")
        return

    # 2) banco de preguntas
    bank = read_sheet("BANCO_PREGUNTAS")
    if bank.empty:
        print("BANCO_PREGUNTAS está vacío. Corre seed_catalogos.py o agrega preguntas.")
        return

    # 3) crear/asegurar evaluación activa
    eval_id = "EVA-001"
    upsert_eval("EVALUACIONES", ["ID_Evaluacion"], {
        "ID_Evaluacion": eval_id,
        "Nombre": "Evaluación Inicial",
        "Fecha": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Estado": "Activa",
        "Descripcion": "Generada por el sistema (v2)"
    })

    # 4) seleccionar activo (simple: el primero; luego lo hacemos desde la GUI)
    asset = inv.iloc[0].to_dict()
    asset_id = str(asset["ID_Activo"])
    tipo = str(asset.get("Tipo_Activo", "General"))

    # parámetro: cuántas preguntas quieres por activo
    n_final = 12

    base_qs = base_questions_for_asset(bank, tipo_activo=tipo, max_n=min(12, len(bank)))
    prompt = build_prompt(asset, base_qs, n_final=n_final)

    print(f"Generando preguntas IA para activo {asset_id} ({tipo}) usando modelo {DEFAULT_MODEL}...")

    text = ollama_generate(DEFAULT_MODEL, prompt)
    parsed = extract_json(text)
    validated = validate_questions(parsed if parsed else [], n_final=n_final)

    # fallback si la IA no devolvió algo válido
    if not validated:
        print("⚠️ IA no devolvió JSON válido. Usando preguntas base sin IA.")
        validated = []
        for i, row in base_qs.head(n_final).iterrows():
            validated.append({
                "ID_Pregunta": str(row["ID_Pregunta"]),
                "Pregunta": str(row["Pregunta"]),
                "Tipo_Respuesta": str(row["Tipo_Respuesta(0/1|1-5)"]),
                "Peso": int(row["Peso(1-5)"]),
                "Dimension": str(row["Dimension(D/I/C)"]),
                "Fuente": "Base"
            })

    # 5) Guardar en Excel: CUESTIONARIO_GENERADO
    fecha = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows_to_write = []
    for q in validated:
        rows_to_write.append({
            "ID_Evaluacion": eval_id,
            "ID_Activo": asset_id,
            "Fecha": fecha,
            "ID_Pregunta": q["ID_Pregunta"],
            "Pregunta": q["Pregunta"],
            "Tipo_Respuesta": q["Tipo_Respuesta"],
            "Peso": q["Peso"],
            "Fuente(IA/Base)": q["Fuente"]
        })

    append_rows("CUESTIONARIO_GENERADO", rows_to_write)
    print(f"✅ Guardadas {len(rows_to_write)} preguntas en CUESTIONARIO_GENERADO para {asset_id} (Eval {eval_id}).")

if __name__ == "__main__":
    main()
