"""
Script para crear catálogos MAGERIT e ISO 27002 en Excel
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_PATH = "matriz_riesgos_v2.xlsx"

print("📚 Creando catálogos MAGERIT e ISO 27002...")

# ============ CRITERIOS MAGERIT (DIC) ============
criterios_magerit = [
    ["Dimensión", "Nivel", "Descripción", "Impacto"],
    # Disponibilidad
    ["D", 1, "Muy Bajo", "Interrupción < 1 hora, sin impacto significativo"],
    ["D", 2, "Bajo", "Interrupción 1-4 horas, impacto operativo menor"],
    ["D", 3, "Medio", "Interrupción 4-24 horas, afecta operaciones normales"],
    ["D", 4, "Alto", "Interrupción 1-3 días, impacto severo en negocio"],
    ["D", 5, "Muy Alto", "Interrupción > 3 días, colapso operacional"],
    # Integridad
    ["I", 1, "Muy Bajo", "Errores menores corregibles sin impacto"],
    ["I", 2, "Bajo", "Errores menores que requieren corrección manual"],
    ["I", 3, "Medio", "Corrupción de datos que afecta operaciones"],
    ["I", 4, "Alto", "Corrupción significativa, pérdida de confiabilidad"],
    ["I", 5, "Muy Alto", "Corrupción masiva, datos irrecuperables"],
    # Confidencialidad
    ["C", 1, "Muy Bajo", "Información pública o de bajo valor"],
    ["C", 2, "Bajo", "Información interna sin datos sensibles"],
    ["C", 3, "Medio", "Información sensible, impacto reputacional"],
    ["C", 4, "Alto", "Información confidencial, impacto legal/financiero"],
    ["C", 5, "Muy Alto", "Información crítica, impacto catastrófico"],
]

# ============ AMENAZAS MAGERIT ============
amenazas_magerit = [
    ["Código", "Categoría", "Amenaza", "Descripción", "Dimensiones", "Severidad"],
    # Desastres naturales
    ["N.1", "Natural", "Fuego", "Incendio que destruye instalaciones o equipos", "D,I", 5],
    ["N.2", "Natural", "Daños por agua", "Inundación, filtración, humedad excesiva", "D,I", 4],
    ["N.*", "Natural", "Desastres naturales", "Terremoto, tormenta eléctrica, etc.", "D,I", 5],
    # Origen industrial
    ["I.1", "Industrial", "Fuego", "Incendio por causas industriales", "D,I", 5],
    ["I.2", "Industrial", "Contaminación", "Polvo, suciedad, agentes corrosivos", "D,I", 3],
    ["I.5", "Industrial", "Avería origen físico", "Desgaste, envejecimiento, fallo de equipo", "D", 4],
    ["I.6", "Industrial", "Corte suministro", "Fallo eléctrico, telecomunicaciones", "D", 5],
    # Errores y fallos no intencionados
    ["E.1", "Humano", "Errores de usuarios", "Operación incorrecta, entrada errónea", "D,I", 3],
    ["E.2", "Humano", "Errores del administrador", "Configuración incorrecta, mantenimiento", "D,I,C", 4],
    ["E.8", "Técnico", "Difusión de software dañino", "Malware, virus, ransomware", "D,I,C", 5],
    ["E.9", "Técnico", "Errores de mantenimiento", "Fallo en proceso de actualización", "D,I", 3],
    ["E.15", "Técnico", "Alteración de secuencia", "Cambio orden de procesos", "I", 3],
    ["E.18", "Técnico", "Destrucción de información", "Borrado accidental", "D,I", 4],
    ["E.19", "Técnico", "Fugas de información", "Divulgación no intencionada", "C", 4],
    ["E.20", "Técnico", "Vulnerabilidades software", "Bugs, backdoors, fallo de seguridad", "D,I,C", 4],
    # Ataques intencionados
    ["A.4", "Ataque", "Manipulación de equipos", "Alteración física maliciosa", "D,I", 4],
    ["A.5", "Ataque", "Suplantación de identidad", "Robo de credenciales, phishing", "I,C", 5],
    ["A.6", "Ataque", "Abuso de privilegios", "Uso indebido de accesos autorizados", "I,C", 4],
    ["A.7", "Ataque", "Uso no previsto", "Uso del sistema fuera de su propósito", "D,I", 3],
    ["A.8", "Ataque", "Difusión de software dañino", "Malware intencional", "D,I,C", 5],
    ["A.9", "Ataque", "Ingeniería social", "Manipulación psicológica", "C", 4],
    ["A.11", "Ataque", "Acceso no autorizado", "Intrusión al sistema", "D,I,C", 5],
    ["A.15", "Ataque", "Modificación deliberada", "Alteración maliciosa de datos", "I", 5],
    ["A.18", "Ataque", "Destrucción de información", "Borrado malicioso", "D,I", 5],
    ["A.19", "Ataque", "Divulgación de información", "Robo de datos, exfiltración", "C", 5],
    ["A.23", "Ataque", "Manipulación de programas", "Inyección código, backdoor", "D,I,C", 5],
    ["A.24", "Ataque", "Denegación de servicio", "DoS, DDoS", "D", 5],
    ["A.25", "Ataque", "Robo", "Sustracción física de equipos/medios", "D,C", 4],
    ["A.26", "Ataque", "Ataque destructivo", "Sabotaje deliberado", "D,I", 5],
    ["A.27", "Ataque", "Ocupación enemiga", "Toma de control físico", "D,I,C", 5],
    ["A.29", "Ataque", "Indisponibilidad del personal", "Huelga, baja, renuncia", "D", 3],
    ["A.30", "Ataque", "Extorsión", "Ransomware, chantaje", "D,C", 5],
]

# ============ CONTROLES ISO 27002:2022 ============
controles_iso27002 = [
    ["Control", "Nombre", "Dominio", "Descripción"],
    # Organizacionales
    ["5.1", "Políticas de seguridad", "Organizacional", "Políticas documentadas y aprobadas"],
    ["5.2", "Roles y responsabilidades", "Organizacional", "Asignación clara de responsabilidades"],
    ["5.7", "Inteligencia de amenazas", "Organizacional", "Monitoreo de amenazas emergentes"],
    ["5.10", "Uso aceptable", "Organizacional", "Políticas de uso correcto de activos"],
    ["5.23", "Seguridad en la nube", "Organizacional", "Controles para servicios cloud"],
    # Personas
    ["6.1", "Investigación de antecedentes", "Personas", "Verificación previa a contratación"],
    ["6.2", "Términos y condiciones", "Personas", "Acuerdos de confidencialidad"],
    ["6.3", "Concienciación", "Personas", "Capacitación en seguridad"],
    ["6.4", "Proceso disciplinario", "Personas", "Sanciones por incumplimiento"],
    ["6.6", "Acuerdos de confidencialidad", "Personas", "NDAs y compromisos"],
    # Físico
    ["7.1", "Perímetros de seguridad física", "Físico", "Barreras y controles de acceso"],
    ["7.2", "Entrada física", "Físico", "Control de acceso a instalaciones"],
    ["7.4", "Monitoreo de seguridad física", "Físico", "Vigilancia, CCTV, alarmas"],
    ["7.7", "Escritorio y pantalla limpios", "Físico", "Clear desk policy"],
    ["7.10", "Medios de almacenamiento", "Físico", "Gestión de discos, USB, backups"],
    ["7.14", "Disposición segura de equipos", "Físico", "Eliminación segura de activos"],
    # Tecnológico
    ["8.1", "Dispositivos de usuario final", "Tecnológico", "Gestión de endpoints"],
    ["8.2", "Derechos de acceso privilegiados", "Tecnológico", "Control de cuentas admin"],
    ["8.3", "Restricción de acceso", "Tecnológico", "Control de acceso lógico"],
    ["8.5", "Autenticación segura", "Tecnológico", "MFA, contraseñas robustas"],
    ["8.8", "Gestión de vulnerabilidades", "Tecnológico", "Escaneo y parcheo"],
    ["8.9", "Gestión de configuración", "Tecnológico", "Hardening, baseline seguro"],
    ["8.10", "Eliminación de información", "Tecnológico", "Borrado seguro de datos"],
    ["8.11", "Enmascaramiento de datos", "Tecnológico", "Protección de datos sensibles"],
    ["8.12", "Prevención de fuga de datos", "Tecnológico", "DLP, controles de exfiltración"],
    ["8.13", "Respaldo de información", "Tecnológico", "Backups regulares y probados"],
    ["8.14", "Redundancia", "Tecnológico", "Alta disponibilidad, clustering"],
    ["8.16", "Actividades de monitoreo", "Tecnológico", "SIEM, logs, alertas"],
    ["8.19", "Instalación de software", "Tecnológico", "Control de aplicaciones"],
    ["8.20", "Seguridad de redes", "Tecnológico", "Firewalls, segmentación"],
    ["8.21", "Seguridad de servicios de red", "Tecnológico", "VPN, VLAN, ACLs"],
    ["8.22", "Segregación de redes", "Tecnológico", "Segmentación por zonas"],
    ["8.23", "Filtrado web", "Tecnológico", "Proxy, categorización de contenido"],
    ["8.24", "Uso de criptografía", "Tecnológico", "Cifrado de datos"],
    ["8.25", "Ciclo de vida de desarrollo", "Tecnológico", "SDLC seguro"],
    ["8.26", "Requisitos de seguridad", "Tecnológico", "Especificación de controles"],
    ["8.28", "Pruebas de seguridad", "Tecnológico", "Pentesting, análisis de código"],
    ["8.31", "Separación de ambientes", "Tecnológico", "Dev, QA, Prod aislados"],
    ["8.34", "Protección contra malware", "Tecnológico", "Antivirus, EDR, sandbox"],
]

# ============ CREAR/ACTUALIZAR EXCEL ============
try:
    wb = load_workbook(EXCEL_PATH)
    
    # Headers style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # ===== CRITERIOS MAGERIT =====
    if "CRITERIOS_MAGERIT" in wb.sheetnames:
        del wb["CRITERIOS_MAGERIT"]
    
    ws_crit = wb.create_sheet("CRITERIOS_MAGERIT")
    for row in criterios_magerit:
        ws_crit.append(row)
    
    # Estilo headers
    for cell in ws_crit[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    print("✅ CRITERIOS_MAGERIT creado")
    
    # ===== AMENAZAS MAGERIT =====
    if "AMENAZAS_MAGERIT" in wb.sheetnames:
        del wb["AMENAZAS_MAGERIT"]
    
    ws_amen = wb.create_sheet("AMENAZAS_MAGERIT")
    for row in amenazas_magerit:
        ws_amen.append(row)
    
    for cell in ws_amen[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    print("✅ AMENAZAS_MAGERIT creado")
    
    # ===== CONTROLES ISO 27002 =====
    if "CONTROLES_ISO27002" in wb.sheetnames:
        del wb["CONTROLES_ISO27002"]
    
    ws_iso = wb.create_sheet("CONTROLES_ISO27002")
    for row in controles_iso27002:
        ws_iso.append(row)
    
    for cell in ws_iso[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    print("✅ CONTROLES_ISO27002 creado")
    
    # Guardar
    wb.save(EXCEL_PATH)
    print(f"\n🎉 Catálogos creados exitosamente en {EXCEL_PATH}")
    print(f"   - {len(criterios_magerit)-1} criterios MAGERIT (DIC)")
    print(f"   - {len(amenazas_magerit)-1} amenazas MAGERIT")
    print(f"   - {len(controles_iso27002)-1} controles ISO 27002:2022")

except Exception as e:
    print(f"❌ Error: {e}")
