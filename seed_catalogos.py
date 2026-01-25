import datetime as dt
from openpyxl import load_workbook

EXCEL_PATH = "matriz_riesgos_v2.xlsx"

def append_rows(ws, rows):
    headers = [c.value for c in ws[1]]
    for r in rows:
        ws.append([r.get(h) for h in headers])

def main():
    wb = load_workbook(EXCEL_PATH)

    # CRITERIOS MAGERIT (plantilla general)
    crit = []
    for dim in ["D", "I", "C"]:
        for lvl in range(1, 6):
            crit.append({
                "Dimension": dim,
                "Nivel(1-5)": lvl,
                "Descripcion": f"Nivel {lvl} para {dim}",
                "Ejemplo": "Se detalla en el documento del proyecto."
            })
    append_rows(wb["CRITERIOS_MAGERIT"], crit)

    # Amenazas MAGERIT (subset inicial)
    amenazas = [
        {"Cod_MAGERIT":"A-01","Categoria":"Acceso","Amenaza":"Acceso no autorizado","Descripcion":"Intentos de acceso indebido",
         "Dimension(D/I/C)":"C","Severidad_Base(1-5)":5},
        {"Cod_MAGERIT":"A-02","Categoria":"Malware","Amenaza":"Infección por malware","Descripcion":"Malware afecta servicios/datos",
         "Dimension(D/I/C)":"I","Severidad_Base(1-5)":4},
        {"Cod_MAGERIT":"A-03","Categoria":"Disponibilidad","Amenaza":"Falla de energía","Descripcion":"Cortes eléctricos",
         "Dimension(D/I/C)":"D","Severidad_Base(1-5)":4},
        {"Cod_MAGERIT":"A-04","Categoria":"Datos","Amenaza":"Pérdida de información","Descripcion":"Borrado/daño de datos",
         "Dimension(D/I/C)":"D","Severidad_Base(1-5)":5},
    ]
    append_rows(wb["CATALOGO_AMENAZAS_MAGERIT"], amenazas)

    # ISO 27002 (subset inicial)
    controles = [
        {"Control":"5.15","Nombre":"Control de acceso","Dominio":"Organizacional","Descripcion":"Gestionar accesos y permisos."},
        {"Control":"8.9","Nombre":"Gestión de configuración","Dominio":"Tecnológico","Descripcion":"Controlar configuraciones base."},
        {"Control":"8.12","Nombre":"Prevención de malware","Dominio":"Tecnológico","Descripcion":"Controles anti-malware."},
        {"Control":"8.13","Nombre":"Copias de seguridad","Dominio":"Tecnológico","Descripcion":"Backups planificados y probados."},
        {"Control":"8.16","Nombre":"Monitoreo de actividades","Dominio":"Tecnológico","Descripcion":"Logs y monitoreo."},
    ]
    append_rows(wb["CATALOGO_ISO27002_2022"], controles)

    # Los bancos de preguntas ahora se crean con crear_bancos_preguntas.py
    # Ya no usamos BANCO_PREGUNTAS genérico

    wb.save(EXCEL_PATH)
    print("Catálogos sembrados (subset inicial).")

if __name__ == "__main__":
    main()

