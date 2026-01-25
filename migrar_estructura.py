"""
Script de migración: Añadir columna ID_Evaluacion a INVENTARIO_ACTIVOS
"""
import pandas as pd
from openpyxl import load_workbook

EXCEL_PATH = "matriz_riesgos_v2.xlsx"

print("🔧 Migrando estructura de INVENTARIO_ACTIVOS...")

try:
    wb = load_workbook(EXCEL_PATH)
    
    if "INVENTARIO_ACTIVOS" in wb.sheetnames:
        ws = wb["INVENTARIO_ACTIVOS"]
        
        # Leer headers actuales
        headers = [cell.value for cell in ws[1]]
        
        # Verificar si ya existe ID_Evaluacion
        if "ID_Evaluacion" not in headers:
            print("➕ Añadiendo columna ID_Evaluacion...")
            
            # Insertar columna al inicio
            ws.insert_cols(1)
            ws.cell(1, 1).value = "ID_Evaluacion"
            
            # Añadir valor por defecto "EVA-001" a todas las filas existentes
            for row in range(2, ws.max_row + 1):
                ws.cell(row, 1).value = "EVA-001"
            
            print("✅ Columna ID_Evaluacion añadida")
        else:
            print("ℹ️ La columna ID_Evaluacion ya existe")
        
        # Verificar otras columnas necesarias
        columnas_necesarias = ["Estado", "Fecha_Creacion", "Descripcion", 
                              "Tipo_Servicio", "App_Critica"]
        
        for col in columnas_necesarias:
            if col not in headers:
                print(f"➕ Añadiendo columna {col}...")
                max_col = ws.max_column + 1
                ws.cell(1, max_col).value = col
                
                # Valores por defecto
                for row in range(2, ws.max_row + 1):
                    if col == "Estado":
                        ws.cell(row, max_col).value = "Pendiente"
                    elif col == "Tipo_Servicio":
                        ws.cell(row, max_col).value = "Otro"
                    elif col == "App_Critica":
                        ws.cell(row, max_col).value = "No"
                    else:
                        ws.cell(row, max_col).value = ""
        
        wb.save(EXCEL_PATH)
        print("✅ Migración completada correctamente")
        
        # Crear evaluación por defecto si no existe
        if "EVALUACIONES" not in wb.sheetnames:
            print("➕ Creando hoja EVALUACIONES...")
            ws_eval = wb.create_sheet("EVALUACIONES")
            ws_eval.append(["ID_Evaluacion", "Nombre", "Descripcion", "Fecha_Creacion", 
                          "Responsable", "Estado", "Origen_Re_Evaluacion"])
            ws_eval.append(["EVA-001", "Evaluación Inicial", "Evaluación migrada desde datos existentes",
                          "2026-01-22", "Sistema", "En Progreso", ""])
            wb.save(EXCEL_PATH)
            print("✅ Evaluación por defecto EVA-001 creada")
    
    else:
        print("⚠️ No existe la hoja INVENTARIO_ACTIVOS")

except Exception as e:
    print(f"❌ Error: {e}")
