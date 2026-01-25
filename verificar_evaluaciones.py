"""
Script para verificar y corregir estructura de EVALUACIONES
"""
import pandas as pd
from openpyxl import load_workbook

EXCEL_PATH = "matriz_riesgos_v2.xlsx"

# Leer hoja actual
try:
    wb = load_workbook(EXCEL_PATH)
    print(f"✅ Excel abierto: {EXCEL_PATH}")
    print(f"📋 Hojas disponibles: {wb.sheetnames}")
    
    if "EVALUACIONES" in wb.sheetnames:
        df = pd.read_excel(EXCEL_PATH, sheet_name="EVALUACIONES")
        print(f"\n📊 Columnas actuales en EVALUACIONES:")
        for i, col in enumerate(df.columns, 1):
            print(f"  {i}. {col}")
        
        print(f"\n📈 Registros: {len(df)}")
        if not df.empty:
            print("\n🔍 Primeros registros:")
            print(df.head())
    else:
        print("\n⚠️ La hoja EVALUACIONES no existe")
    
    wb.close()
    
except Exception as e:
    print(f"❌ Error: {str(e)}")
