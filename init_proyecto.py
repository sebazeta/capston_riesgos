"""
Script de inicialización del Proyecto TITA
Ejecuta este script para configurar todo desde cero

Uso: python init_proyecto.py
"""
import os
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

EXCEL_PATH = "matriz_riesgos_v2.xlsx"
BACKUP_PATH = "matriz_riesgos_v2_backup.xlsx"

# ==================== ESTRUCTURA DE LA BASE DE DATOS ====================
SHEETS = {
    # Metadatos
    "PORTADA": ["Campo", "Valor"],
    "EVALUACIONES": ["ID_Evaluacion", "Nombre", "Fecha", "Estado", "Descripcion"],

    # Catálogos de referencia
    "CRITERIOS_MAGERIT": ["Dimension", "Nivel(1-5)", "Descripcion", "Ejemplo"],
    "CATALOGO_AMENAZAS_MAGERIT": ["Cod_MAGERIT", "Categoria", "Amenaza", "Descripcion", "Dimension(D/I/C)", "Severidad_Base(1-5)"],
    "CATALOGO_ISO27002_2022": ["Control", "Nombre", "Dominio", "Descripcion"],

    # Inventario de activos - columnas sincronizadas con activo_service.py
    "INVENTARIO_ACTIVOS": ["ID_Evaluacion", "ID_Activo", "Nombre_Activo", "Tipo_Activo", 
                           "Ubicacion", "Propietario", "Tipo_Servicio", "App_Critica",
                           "Estado", "Fecha_Creacion"],

    # Bancos de preguntas por tipo de activo
    "BANCO_PREGUNTAS_FISICAS": ["ID_Pregunta", "Tipo_Activo", "Bloque", "Dimension", "Pregunta", 
                                 "Opcion_1", "Opcion_2", "Opcion_3", "Opcion_4", "Peso"],
    "BANCO_PREGUNTAS_VIRTUALES": ["ID_Pregunta", "Tipo_Activo", "Bloque", "Dimension", "Pregunta", 
                                   "Opcion_1", "Opcion_2", "Opcion_3", "Opcion_4", "Peso"],

    # Cuestionarios generados para cada activo
    "CUESTIONARIOS": ["ID_Evaluacion", "ID_Activo", "Fecha_Version", "ID_Pregunta", "Bloque",
                      "Dimension", "Pregunta", "Opcion_1", "Opcion_2", "Opcion_3", "Opcion_4", "Peso"],

    # Respuestas del usuario
    "RESPUESTAS": ["ID_Evaluacion", "ID_Activo", "Fecha_Cuestionario", "ID_Pregunta", "Bloque",
                   "Pregunta", "Respuesta", "Valor_Numerico", "Peso", "Dimension", "Fecha"],

    # Impacto DIC (valoración MAGERIT)
    "IMPACTO_ACTIVOS": ["ID_Evaluacion", "ID_Activo", "Fecha", "Impacto_D", "Impacto_I", "Impacto_C",
                        "Justificacion_D", "Justificacion_I", "Justificacion_C"],

    # Análisis de riesgos
    "AMENAZAS_VULNERAB": ["ID_Evaluacion", "ID_Riesgo", "ID_Activo", "Fecha", "Cod_MAGERIT", "Amenaza",
                          "Vulnerabilidad", "Dimension(D/I/C)", "Severidad(1-5)"],

    "ANALISIS_RIESGO": ["ID_Evaluacion", "ID_Activo", "Fecha", "Tipo_Activo", "Nombre_Activo",
                        "Probabilidad", "Impacto", "Riesgo_Inherente", "Nivel_Riesgo",
                        "Recomendaciones", "Estado", "Modelo_IA"],

    "SALVAGUARDAS": ["ID_Evaluacion", "ID_Riesgo", "ID_Activo", "Fecha", "Control_ISO27002", 
                     "Nombre_Control", "Descripcion", "Efectividad(0-100)"],

    "RIESGO_RESIDUAL": ["ID_Evaluacion", "ID_Riesgo", "ID_Activo", "Fecha", 
                        "Riesgo_Inherente", "Efectividad", "Riesgo_Residual"],

    # Métricas y reportes
    "HISTORICO": ["ID_Evaluacion", "Fecha", "Total_Activos", "Riesgo_Promedio_Residual", 
                  "Riesgos_Altos", "Riesgos_Medios", "Riesgos_Bajos"],
    "DASHBOARD": ["KPI", "Valor"],
    "COMPARATIVO": ["Evaluacion_A", "Evaluacion_B", "Metrica", "Valor_A", "Valor_B", "Diferencia"],
}


# ==================== CATÁLOGOS INICIALES ====================
def get_criterios_magerit():
    """Criterios MAGERIT para valoración DIC"""
    return [
        {"Dimension": "D", "Nivel(1-5)": 1, "Descripcion": "Impacto mínimo en disponibilidad", "Ejemplo": "Interrupción < 1 hora"},
        {"Dimension": "D", "Nivel(1-5)": 2, "Descripcion": "Impacto bajo en disponibilidad", "Ejemplo": "Interrupción 1-4 horas"},
        {"Dimension": "D", "Nivel(1-5)": 3, "Descripcion": "Impacto moderado en disponibilidad", "Ejemplo": "Interrupción 4-24 horas"},
        {"Dimension": "D", "Nivel(1-5)": 4, "Descripcion": "Impacto alto en disponibilidad", "Ejemplo": "Interrupción 1-7 días"},
        {"Dimension": "D", "Nivel(1-5)": 5, "Descripcion": "Impacto crítico en disponibilidad", "Ejemplo": "Interrupción > 7 días"},
        {"Dimension": "I", "Nivel(1-5)": 1, "Descripcion": "Impacto mínimo en integridad", "Ejemplo": "Errores menores sin consecuencia"},
        {"Dimension": "I", "Nivel(1-5)": 2, "Descripcion": "Impacto bajo en integridad", "Ejemplo": "Errores detectables y corregibles"},
        {"Dimension": "I", "Nivel(1-5)": 3, "Descripcion": "Impacto moderado en integridad", "Ejemplo": "Datos incorrectos con impacto operativo"},
        {"Dimension": "I", "Nivel(1-5)": 4, "Descripcion": "Impacto alto en integridad", "Ejemplo": "Corrupción de datos críticos"},
        {"Dimension": "I", "Nivel(1-5)": 5, "Descripcion": "Impacto crítico en integridad", "Ejemplo": "Pérdida total de integridad de datos"},
        {"Dimension": "C", "Nivel(1-5)": 1, "Descripcion": "Impacto mínimo en confidencialidad", "Ejemplo": "Información pública"},
        {"Dimension": "C", "Nivel(1-5)": 2, "Descripcion": "Impacto bajo en confidencialidad", "Ejemplo": "Información interna no sensible"},
        {"Dimension": "C", "Nivel(1-5)": 3, "Descripcion": "Impacto moderado en confidencialidad", "Ejemplo": "Información sensible de la organización"},
        {"Dimension": "C", "Nivel(1-5)": 4, "Descripcion": "Impacto alto en confidencialidad", "Ejemplo": "Datos personales o confidenciales"},
        {"Dimension": "C", "Nivel(1-5)": 5, "Descripcion": "Impacto crítico en confidencialidad", "Ejemplo": "Secretos comerciales o datos críticos"},
    ]


def get_amenazas_magerit():
    """Catálogo de amenazas MAGERIT"""
    return [
        {"Cod_MAGERIT": "N.1", "Categoria": "Desastres naturales", "Amenaza": "Fuego", "Descripcion": "Incendio en instalaciones", "Dimension(D/I/C)": "D", "Severidad_Base(1-5)": 5},
        {"Cod_MAGERIT": "N.2", "Categoria": "Desastres naturales", "Amenaza": "Daños por agua", "Descripcion": "Inundaciones, fugas", "Dimension(D/I/C)": "D", "Severidad_Base(1-5)": 4},
        {"Cod_MAGERIT": "I.5", "Categoria": "Industrial", "Amenaza": "Avería de origen físico", "Descripcion": "Fallo de hardware", "Dimension(D/I/C)": "D", "Severidad_Base(1-5)": 3},
        {"Cod_MAGERIT": "I.6", "Categoria": "Industrial", "Amenaza": "Corte de suministro eléctrico", "Descripcion": "Fallo en alimentación", "Dimension(D/I/C)": "D", "Severidad_Base(1-5)": 4},
        {"Cod_MAGERIT": "E.1", "Categoria": "Errores", "Amenaza": "Errores de usuarios", "Descripcion": "Errores humanos en operación", "Dimension(D/I/C)": "I", "Severidad_Base(1-5)": 2},
        {"Cod_MAGERIT": "E.2", "Categoria": "Errores", "Amenaza": "Errores de administrador", "Descripcion": "Errores en configuración", "Dimension(D/I/C)": "I", "Severidad_Base(1-5)": 3},
        {"Cod_MAGERIT": "E.8", "Categoria": "Errores", "Amenaza": "Difusión de software dañino", "Descripcion": "Virus, malware", "Dimension(D/I/C)": "D", "Severidad_Base(1-5)": 4},
        {"Cod_MAGERIT": "A.5", "Categoria": "Ataques", "Amenaza": "Suplantación de identidad", "Descripcion": "Phishing, ingeniería social", "Dimension(D/I/C)": "C", "Severidad_Base(1-5)": 4},
        {"Cod_MAGERIT": "A.11", "Categoria": "Ataques", "Amenaza": "Acceso no autorizado", "Descripcion": "Intrusión en sistemas", "Dimension(D/I/C)": "C", "Severidad_Base(1-5)": 5},
        {"Cod_MAGERIT": "A.24", "Categoria": "Ataques", "Amenaza": "Denegación de servicio", "Descripcion": "DoS/DDoS", "Dimension(D/I/C)": "D", "Severidad_Base(1-5)": 4},
    ]


def get_controles_iso27002():
    """Controles ISO 27002:2022"""
    return [
        {"Control": "5.15", "Nombre": "Control de acceso", "Dominio": "Organizacional", "Descripcion": "Gestionar accesos y permisos."},
        {"Control": "8.9", "Nombre": "Gestión de configuración", "Dominio": "Tecnológico", "Descripcion": "Controlar configuraciones base."},
        {"Control": "8.12", "Nombre": "Prevención de malware", "Dominio": "Tecnológico", "Descripcion": "Controles anti-malware."},
        {"Control": "8.13", "Nombre": "Copias de seguridad", "Dominio": "Tecnológico", "Descripcion": "Backups planificados y probados."},
        {"Control": "8.16", "Nombre": "Monitoreo de actividades", "Dominio": "Tecnológico", "Descripcion": "Logs y monitoreo."},
    ]


# ==================== BANCOS DE PREGUNTAS OFICIALES ====================
def get_banco_preguntas_fisicas():
    """Banco de 21 preguntas para Servidores Físicos"""
    return [
        # Bloque A: Impacto (5 preguntas)
        {"ID_Pregunta": "PF-A01", "Tipo_Activo": "Servidor Físico", "Bloque": "A-Impacto", "Dimension": "D",
         "Pregunta": "¿Cuál es el tiempo máximo tolerable de interrupción (RTO) del servidor?",
         "Opcion_1": "Más de 72 horas", "Opcion_2": "24-72 horas", "Opcion_3": "4-24 horas", "Opcion_4": "Menos de 4 horas", "Peso": 5},
        {"ID_Pregunta": "PF-A02", "Tipo_Activo": "Servidor Físico", "Bloque": "A-Impacto", "Dimension": "D",
         "Pregunta": "¿Cuántos usuarios o procesos críticos dependen directamente de este servidor?",
         "Opcion_1": "Menos de 10", "Opcion_2": "10-50", "Opcion_3": "50-200", "Opcion_4": "Más de 200", "Peso": 4},
        {"ID_Pregunta": "PF-A03", "Tipo_Activo": "Servidor Físico", "Bloque": "A-Impacto", "Dimension": "I",
         "Pregunta": "¿Qué nivel de pérdida de datos es tolerable (RPO)?",
         "Opcion_1": "Hasta 1 semana", "Opcion_2": "Hasta 24 horas", "Opcion_3": "Hasta 4 horas", "Opcion_4": "Cero pérdida", "Peso": 5},
        {"ID_Pregunta": "PF-A04", "Tipo_Activo": "Servidor Físico", "Bloque": "A-Impacto", "Dimension": "C",
         "Pregunta": "¿Qué tipo de información procesa este servidor?",
         "Opcion_1": "Pública", "Opcion_2": "Interna", "Opcion_3": "Confidencial", "Opcion_4": "Altamente sensible", "Peso": 5},
        {"ID_Pregunta": "PF-A05", "Tipo_Activo": "Servidor Físico", "Bloque": "A-Impacto", "Dimension": "D",
         "Pregunta": "¿Cuál sería el impacto financiero estimado por hora de inactividad?",
         "Opcion_1": "Menor a $100", "Opcion_2": "$100-$1,000", "Opcion_3": "$1,000-$10,000", "Opcion_4": "Mayor a $10,000", "Peso": 4},
        
        # Bloque B: Continuidad (4 preguntas)
        {"ID_Pregunta": "PF-B01", "Tipo_Activo": "Servidor Físico", "Bloque": "B-Continuidad", "Dimension": "D",
         "Pregunta": "¿Existe un servidor de respaldo o failover configurado?",
         "Opcion_1": "No existe", "Opcion_2": "Existe sin probar", "Opcion_3": "Probado anualmente", "Opcion_4": "Activo-Activo probado", "Peso": 5},
        {"ID_Pregunta": "PF-B02", "Tipo_Activo": "Servidor Físico", "Bloque": "B-Continuidad", "Dimension": "D",
         "Pregunta": "¿Con qué frecuencia se realizan copias de seguridad?",
         "Opcion_1": "Mensual o nunca", "Opcion_2": "Semanal", "Opcion_3": "Diario", "Opcion_4": "Continuo/Tiempo real", "Peso": 5},
        {"ID_Pregunta": "PF-B03", "Tipo_Activo": "Servidor Físico", "Bloque": "B-Continuidad", "Dimension": "D",
         "Pregunta": "¿Se prueban regularmente las restauraciones de backup?",
         "Opcion_1": "Nunca", "Opcion_2": "Anualmente", "Opcion_3": "Trimestralmente", "Opcion_4": "Mensualmente", "Peso": 4},
        {"ID_Pregunta": "PF-B04", "Tipo_Activo": "Servidor Físico", "Bloque": "B-Continuidad", "Dimension": "D",
         "Pregunta": "¿El servidor tiene fuente de alimentación redundante (UPS)?",
         "Opcion_1": "Sin UPS", "Opcion_2": "UPS básico", "Opcion_3": "UPS + generador", "Opcion_4": "Redundancia completa", "Peso": 3},
        
        # Bloque C: Controles (5 preguntas)
        {"ID_Pregunta": "PF-C01", "Tipo_Activo": "Servidor Físico", "Bloque": "C-Controles", "Dimension": "C",
         "Pregunta": "¿Cómo se gestiona el control de acceso al servidor?",
         "Opcion_1": "Sin control", "Opcion_2": "Usuario/contraseña básico", "Opcion_3": "Roles y permisos", "Opcion_4": "MFA + roles + auditoría", "Peso": 5},
        {"ID_Pregunta": "PF-C02", "Tipo_Activo": "Servidor Físico", "Bloque": "C-Controles", "Dimension": "I",
         "Pregunta": "¿Con qué frecuencia se aplican parches de seguridad?",
         "Opcion_1": "Nunca/Raramente", "Opcion_2": "Anualmente", "Opcion_3": "Trimestralmente", "Opcion_4": "Mensual o automático", "Peso": 5},
        {"ID_Pregunta": "PF-C03", "Tipo_Activo": "Servidor Físico", "Bloque": "C-Controles", "Dimension": "D",
         "Pregunta": "¿Existe monitoreo de rendimiento y alertas?",
         "Opcion_1": "Sin monitoreo", "Opcion_2": "Monitoreo manual", "Opcion_3": "Alertas básicas", "Opcion_4": "Monitoreo 24/7 con escalamiento", "Peso": 4},
        {"ID_Pregunta": "PF-C04", "Tipo_Activo": "Servidor Físico", "Bloque": "C-Controles", "Dimension": "C",
         "Pregunta": "¿Se registran y revisan los logs de acceso?",
         "Opcion_1": "Sin logs", "Opcion_2": "Logs sin revisión", "Opcion_3": "Revisión mensual", "Opcion_4": "SIEM con alertas", "Peso": 4},
        {"ID_Pregunta": "PF-C05", "Tipo_Activo": "Servidor Físico", "Bloque": "C-Controles", "Dimension": "C",
         "Pregunta": "¿Existe segmentación de red para este servidor?",
         "Opcion_1": "Red plana", "Opcion_2": "VLAN básica", "Opcion_3": "Firewall dedicado", "Opcion_4": "Microsegmentación", "Peso": 4},
        
        # Bloque D: Ciberseguridad (4 preguntas)
        {"ID_Pregunta": "PF-D01", "Tipo_Activo": "Servidor Físico", "Bloque": "D-Ciberseguridad", "Dimension": "I",
         "Pregunta": "¿El servidor tiene antivirus/antimalware actualizado?",
         "Opcion_1": "Sin protección", "Opcion_2": "Antivirus básico", "Opcion_3": "EDR", "Opcion_4": "EDR + XDR integrado", "Peso": 4},
        {"ID_Pregunta": "PF-D02", "Tipo_Activo": "Servidor Físico", "Bloque": "D-Ciberseguridad", "Dimension": "C",
         "Pregunta": "¿Los datos en reposo están cifrados?",
         "Opcion_1": "Sin cifrado", "Opcion_2": "Cifrado parcial", "Opcion_3": "Cifrado completo", "Opcion_4": "Cifrado + gestión de claves", "Peso": 4},
        {"ID_Pregunta": "PF-D03", "Tipo_Activo": "Servidor Físico", "Bloque": "D-Ciberseguridad", "Dimension": "I",
         "Pregunta": "¿Se realizan análisis de vulnerabilidades?",
         "Opcion_1": "Nunca", "Opcion_2": "Anualmente", "Opcion_3": "Trimestralmente", "Opcion_4": "Continuo automatizado", "Peso": 4},
        {"ID_Pregunta": "PF-D04", "Tipo_Activo": "Servidor Físico", "Bloque": "D-Ciberseguridad", "Dimension": "D",
         "Pregunta": "¿Existe protección contra ransomware específica?",
         "Opcion_1": "Sin protección", "Opcion_2": "Backups offline", "Opcion_3": "Backups + detección", "Opcion_4": "Protección multicapa", "Peso": 5},
        
        # Bloque E: Exposición (3 preguntas)
        {"ID_Pregunta": "PF-E01", "Tipo_Activo": "Servidor Físico", "Bloque": "E-Exposición", "Dimension": "C",
         "Pregunta": "¿El servidor tiene servicios expuestos a Internet?",
         "Opcion_1": "Totalmente expuesto", "Opcion_2": "Parcialmente expuesto", "Opcion_3": "Solo VPN", "Opcion_4": "Solo red interna", "Peso": 5},
        {"ID_Pregunta": "PF-E02", "Tipo_Activo": "Servidor Físico", "Bloque": "E-Exposición", "Dimension": "C",
         "Pregunta": "¿Cuál es el nivel de acceso físico al servidor?",
         "Opcion_1": "Acceso libre", "Opcion_2": "Sala cerrada", "Opcion_3": "Datacenter con control", "Opcion_4": "Datacenter Tier III+", "Peso": 3},
        {"ID_Pregunta": "PF-E03", "Tipo_Activo": "Servidor Físico", "Bloque": "E-Exposición", "Dimension": "D",
         "Pregunta": "¿Cuántas dependencias externas tiene el servidor?",
         "Opcion_1": "Más de 10", "Opcion_2": "5-10", "Opcion_3": "2-4", "Opcion_4": "0-1", "Peso": 3},
    ]


def get_banco_preguntas_virtuales():
    """Banco de 21 preguntas para Servidores Virtuales"""
    return [
        # Bloque A: Impacto (5 preguntas)
        {"ID_Pregunta": "PV-A01", "Tipo_Activo": "Servidor Virtual", "Bloque": "A-Impacto", "Dimension": "D",
         "Pregunta": "¿Cuál es el tiempo máximo tolerable de interrupción (RTO) de la VM?",
         "Opcion_1": "Más de 72 horas", "Opcion_2": "24-72 horas", "Opcion_3": "4-24 horas", "Opcion_4": "Menos de 4 horas", "Peso": 5},
        {"ID_Pregunta": "PV-A02", "Tipo_Activo": "Servidor Virtual", "Bloque": "A-Impacto", "Dimension": "D",
         "Pregunta": "¿Cuántos servicios o aplicaciones dependen de esta VM?",
         "Opcion_1": "1-2 servicios", "Opcion_2": "3-5 servicios", "Opcion_3": "6-10 servicios", "Opcion_4": "Más de 10 servicios", "Peso": 4},
        {"ID_Pregunta": "PV-A03", "Tipo_Activo": "Servidor Virtual", "Bloque": "A-Impacto", "Dimension": "I",
         "Pregunta": "¿Qué nivel de pérdida de datos es tolerable (RPO)?",
         "Opcion_1": "Hasta 1 semana", "Opcion_2": "Hasta 24 horas", "Opcion_3": "Hasta 4 horas", "Opcion_4": "Cero pérdida", "Peso": 5},
        {"ID_Pregunta": "PV-A04", "Tipo_Activo": "Servidor Virtual", "Bloque": "A-Impacto", "Dimension": "C",
         "Pregunta": "¿Qué tipo de información procesa esta VM?",
         "Opcion_1": "Pública", "Opcion_2": "Interna", "Opcion_3": "Confidencial", "Opcion_4": "Altamente sensible", "Peso": 5},
        {"ID_Pregunta": "PV-A05", "Tipo_Activo": "Servidor Virtual", "Bloque": "A-Impacto", "Dimension": "D",
         "Pregunta": "¿Esta VM forma parte de un cluster o granja de servidores?",
         "Opcion_1": "VM aislada crítica", "Opcion_2": "VM aislada no crítica", "Opcion_3": "Parte de cluster", "Opcion_4": "Cluster con auto-scaling", "Peso": 4},
        
        # Bloque B: Continuidad (4 preguntas)
        {"ID_Pregunta": "PV-B01", "Tipo_Activo": "Servidor Virtual", "Bloque": "B-Continuidad", "Dimension": "D",
         "Pregunta": "¿Existe capacidad de migración en vivo (vMotion/Live Migration)?",
         "Opcion_1": "No disponible", "Opcion_2": "Disponible sin probar", "Opcion_3": "Probado anualmente", "Opcion_4": "Automatizado DRS/HA", "Peso": 5},
        {"ID_Pregunta": "PV-B02", "Tipo_Activo": "Servidor Virtual", "Bloque": "B-Continuidad", "Dimension": "D",
         "Pregunta": "¿Con qué frecuencia se realizan snapshots/backups de la VM?",
         "Opcion_1": "Mensual o nunca", "Opcion_2": "Semanal", "Opcion_3": "Diario", "Opcion_4": "Múltiples veces al día", "Peso": 5},
        {"ID_Pregunta": "PV-B03", "Tipo_Activo": "Servidor Virtual", "Bloque": "B-Continuidad", "Dimension": "D",
         "Pregunta": "¿Existe réplica de la VM en otro sitio/datacenter?",
         "Opcion_1": "Sin réplica", "Opcion_2": "Réplica manual", "Opcion_3": "Réplica asíncrona", "Opcion_4": "Réplica síncrona multi-sitio", "Peso": 4},
        {"ID_Pregunta": "PV-B04", "Tipo_Activo": "Servidor Virtual", "Bloque": "B-Continuidad", "Dimension": "D",
         "Pregunta": "¿El hypervisor tiene recursos reservados para esta VM?",
         "Opcion_1": "Sin reservas", "Opcion_2": "Reserva parcial", "Opcion_3": "Reserva completa", "Opcion_4": "Host dedicado", "Peso": 3},
        
        # Bloque C: Controles (5 preguntas)
        {"ID_Pregunta": "PV-C01", "Tipo_Activo": "Servidor Virtual", "Bloque": "C-Controles", "Dimension": "C",
         "Pregunta": "¿Cómo se gestiona el acceso a la consola de la VM?",
         "Opcion_1": "Sin control", "Opcion_2": "Usuario/contraseña", "Opcion_3": "Roles + auditoría", "Opcion_4": "PAM + MFA + grabación", "Peso": 5},
        {"ID_Pregunta": "PV-C02", "Tipo_Activo": "Servidor Virtual", "Bloque": "C-Controles", "Dimension": "I",
         "Pregunta": "¿La imagen/template de la VM está hardened?",
         "Opcion_1": "Instalación default", "Opcion_2": "Configuración básica", "Opcion_3": "CIS Benchmark parcial", "Opcion_4": "CIS Benchmark completo", "Peso": 4},
        {"ID_Pregunta": "PV-C03", "Tipo_Activo": "Servidor Virtual", "Bloque": "C-Controles", "Dimension": "D",
         "Pregunta": "¿Existe monitoreo de recursos de la VM (CPU, RAM, disco)?",
         "Opcion_1": "Sin monitoreo", "Opcion_2": "Monitoreo básico", "Opcion_3": "Alertas automáticas", "Opcion_4": "AIOps con predicción", "Peso": 4},
        {"ID_Pregunta": "PV-C04", "Tipo_Activo": "Servidor Virtual", "Bloque": "C-Controles", "Dimension": "I",
         "Pregunta": "¿Con qué frecuencia se actualiza el SO de la VM?",
         "Opcion_1": "Nunca/Raramente", "Opcion_2": "Anualmente", "Opcion_3": "Trimestralmente", "Opcion_4": "Mensual automatizado", "Peso": 5},
        {"ID_Pregunta": "PV-C05", "Tipo_Activo": "Servidor Virtual", "Bloque": "C-Controles", "Dimension": "C",
         "Pregunta": "¿Se utilizan políticas de grupo o configuración centralizada?",
         "Opcion_1": "Configuración manual", "Opcion_2": "Scripts básicos", "Opcion_3": "GPO/Ansible parcial", "Opcion_4": "IaC completo (Terraform/Ansible)", "Peso": 3},
        
        # Bloque D: Ciberseguridad (4 preguntas)
        {"ID_Pregunta": "PV-D01", "Tipo_Activo": "Servidor Virtual", "Bloque": "D-Ciberseguridad", "Dimension": "I",
         "Pregunta": "¿La VM tiene agente de seguridad endpoint (EDR/XDR)?",
         "Opcion_1": "Sin protección", "Opcion_2": "Antivirus básico", "Opcion_3": "EDR", "Opcion_4": "XDR integrado con SOAR", "Peso": 4},
        {"ID_Pregunta": "PV-D02", "Tipo_Activo": "Servidor Virtual", "Bloque": "D-Ciberseguridad", "Dimension": "C",
         "Pregunta": "¿Los discos virtuales están cifrados?",
         "Opcion_1": "Sin cifrado", "Opcion_2": "Cifrado storage", "Opcion_3": "Cifrado VM individual", "Opcion_4": "Cifrado + vTPM + SecureBoot", "Peso": 4},
        {"ID_Pregunta": "PV-D03", "Tipo_Activo": "Servidor Virtual", "Bloque": "D-Ciberseguridad", "Dimension": "C",
         "Pregunta": "¿Existe segmentación de red virtual (NSX/micro-segmentación)?",
         "Opcion_1": "Sin segmentación", "Opcion_2": "VLANs básicas", "Opcion_3": "Firewall distribuido", "Opcion_4": "Zero Trust NSX-T", "Peso": 4},
        {"ID_Pregunta": "PV-D04", "Tipo_Activo": "Servidor Virtual", "Bloque": "D-Ciberseguridad", "Dimension": "I",
         "Pregunta": "¿Se monitorea la integridad de archivos del sistema?",
         "Opcion_1": "Sin monitoreo", "Opcion_2": "Verificación manual", "Opcion_3": "FIM básico", "Opcion_4": "FIM + respuesta automática", "Peso": 3},
        
        # Bloque E: Exposición (3 preguntas)
        {"ID_Pregunta": "PV-E01", "Tipo_Activo": "Servidor Virtual", "Bloque": "E-Exposición", "Dimension": "C",
         "Pregunta": "¿La VM está en nube pública, privada o híbrida?",
         "Opcion_1": "Nube pública sin controles", "Opcion_2": "Nube pública con controles", "Opcion_3": "Nube privada", "Opcion_4": "On-premise aislado", "Peso": 4},
        {"ID_Pregunta": "PV-E02", "Tipo_Activo": "Servidor Virtual", "Bloque": "E-Exposición", "Dimension": "C",
         "Pregunta": "¿La VM tiene interfaces de red expuestas a Internet?",
         "Opcion_1": "IP pública directa", "Opcion_2": "NAT con puertos abiertos", "Opcion_3": "Solo a través de LB/WAF", "Opcion_4": "Solo red privada", "Peso": 5},
        {"ID_Pregunta": "PV-E03", "Tipo_Activo": "Servidor Virtual", "Bloque": "E-Exposición", "Dimension": "D",
         "Pregunta": "¿Cuántas VMs comparten el mismo host físico?",
         "Opcion_1": "Más de 50", "Opcion_2": "20-50", "Opcion_3": "5-20", "Opcion_4": "Host dedicado", "Peso": 3},
    ]


# ==================== FUNCIONES DE CREACIÓN ====================
def autosize(ws):
    """Ajusta el ancho de las columnas automáticamente"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def append_rows_to_sheet(ws, headers, rows):
    """Añade filas a una hoja"""
    for row in rows:
        ws.append([row.get(h) for h in headers])


def crear_excel():
    """Crea el archivo Excel con toda la estructura"""
    print("🔧 Creando archivo Excel...")
    
    # Eliminar archivos existentes
    for path in [EXCEL_PATH, BACKUP_PATH]:
        if os.path.exists(path):
            os.unlink(path)
            print(f"   Eliminado: {path}")
    
    # Crear workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Crear todas las hojas
    for name, headers in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        autosize(ws)
    print(f"   ✅ Creadas {len(SHEETS)} hojas")
    
    # Poblar catálogos
    print("📚 Poblando catálogos...")
    
    # CRITERIOS_MAGERIT
    ws = wb["CRITERIOS_MAGERIT"]
    headers = SHEETS["CRITERIOS_MAGERIT"]
    append_rows_to_sheet(ws, headers, get_criterios_magerit())
    print("   ✅ CRITERIOS_MAGERIT")
    
    # CATALOGO_AMENAZAS_MAGERIT
    ws = wb["CATALOGO_AMENAZAS_MAGERIT"]
    headers = SHEETS["CATALOGO_AMENAZAS_MAGERIT"]
    append_rows_to_sheet(ws, headers, get_amenazas_magerit())
    print("   ✅ CATALOGO_AMENAZAS_MAGERIT")
    
    # CATALOGO_ISO27002_2022
    ws = wb["CATALOGO_ISO27002_2022"]
    headers = SHEETS["CATALOGO_ISO27002_2022"]
    append_rows_to_sheet(ws, headers, get_controles_iso27002())
    print("   ✅ CATALOGO_ISO27002_2022")
    
    # Bancos de preguntas
    print("📝 Creando bancos de preguntas...")
    
    # BANCO_PREGUNTAS_FISICAS
    ws = wb["BANCO_PREGUNTAS_FISICAS"]
    headers = SHEETS["BANCO_PREGUNTAS_FISICAS"]
    preguntas_fisicas = get_banco_preguntas_fisicas()
    append_rows_to_sheet(ws, headers, preguntas_fisicas)
    print(f"   ✅ BANCO_PREGUNTAS_FISICAS ({len(preguntas_fisicas)} preguntas)")
    
    # BANCO_PREGUNTAS_VIRTUALES
    ws = wb["BANCO_PREGUNTAS_VIRTUALES"]
    headers = SHEETS["BANCO_PREGUNTAS_VIRTUALES"]
    preguntas_virtuales = get_banco_preguntas_virtuales()
    append_rows_to_sheet(ws, headers, preguntas_virtuales)
    print(f"   ✅ BANCO_PREGUNTAS_VIRTUALES ({len(preguntas_virtuales)} preguntas)")
    
    # Guardar
    wb.save(EXCEL_PATH)
    wb.close()
    
    # Crear backup inicial
    import shutil
    shutil.copy2(EXCEL_PATH, BACKUP_PATH)
    
    print(f"\n🎉 Proyecto inicializado correctamente!")
    print(f"   📄 Base de datos: {EXCEL_PATH}")
    print(f"   💾 Backup: {BACKUP_PATH}")
    print(f"\n📋 Resumen:")
    print(f"   - {len(SHEETS)} hojas creadas")
    print(f"   - 15 criterios MAGERIT")
    print(f"   - 10 amenazas catalogadas")
    print(f"   - 5 controles ISO 27002")
    print(f"   - {len(preguntas_fisicas)} preguntas para Servidores Físicos")
    print(f"   - {len(preguntas_virtuales)} preguntas para Servidores Virtuales")
    print(f"\n🚀 Para iniciar la aplicación:")
    print(f"   streamlit run app_final.py --server.port 8506")


if __name__ == "__main__":
    print("=" * 60)
    print("🏢 PROYECTO TITA - Inicialización")
    print("=" * 60)
    crear_excel()
