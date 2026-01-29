"""
Servicio para el Modelo de Matriz MAGERIT
Replica exactamente la estructura de la matriz Excel de referencia:
- CRITERIOS DE VALORACIÓN
- ACTIVOS
- IDENTIFICACION_VALORACION
- VULNERABILIDADES_AMENAZAS
- RIESGO
- MAPA_RIESGOS
- RIESGO_ACTIVOS
- SALVAGUARDAS

Escalas:
- D/I/C: 0-3 (Nula, Baja, Media, Alta)
- Frecuencia: 0.1-3 (Cada varios años → Diario)
- Criticidad: MAX(D, I, C)
- Impacto: Criticidad × MAX(Deg_D, Deg_I, Deg_C)
- Riesgo: Frecuencia × Impacto
"""
import sqlite3
import pandas as pd
import datetime as dt
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass
from services.database_service import get_connection, DB_PATH

# ==================== CONSTANTES (ESCALAS) ====================

ESCALA_DISPONIBILIDAD = [
    {"nivel": "Nula", "valor": 0, "descripcion": "Inaccesibilidad no afecta actividad normal"},
    {"nivel": "Baja", "valor": 1, "descripcion": "Inaccesibilidad de 1 semana ocasiona perjuicio menor"},
    {"nivel": "Media", "valor": 2, "descripcion": "Inaccesibilidad de 1 jornada impide actividades"},
    {"nivel": "Alta", "valor": 3, "descripcion": "Inaccesibilidad de 1 hora impide actividades críticas"},
]

ESCALA_INTEGRIDAD = [
    {"nivel": "Nula", "valor": 0, "descripcion": "Modificación reparable fácilmente, sin afectación"},
    {"nivel": "Baja", "valor": 1, "descripcion": "Modificación reparable, perjuicio menor"},
    {"nivel": "Media", "valor": 2, "descripcion": "Modificación difícil de reparar, perjuicio significativo"},
    {"nivel": "Alta", "valor": 3, "descripcion": "Modificación no autorizada no puede repararse"},
]

ESCALA_CONFIDENCIALIDAD = [
    {"nivel": "Nula", "valor": 0, "descripcion": "Cualquier persona dentro o fuera de la empresa"},
    {"nivel": "Baja", "valor": 1, "descripcion": "Todos los empleados de la empresa"},
    {"nivel": "Media", "valor": 2, "descripcion": "Solo quienes necesitan para su trabajo"},
    {"nivel": "Alta", "valor": 3, "descripcion": "Solo grupo muy reducido, divulgación = perjuicio grave"},
]

ESCALA_CRITICIDAD = [
    {"nivel": "Nula", "valor": 0, "criterio": "Si todos D, I, C son 0"},
    {"nivel": "Baja", "valor": 1, "criterio": "Si MAX(D,I,C) = 1"},
    {"nivel": "Media", "valor": 2, "criterio": "Si MAX(D,I,C) = 2"},
    {"nivel": "Alta", "valor": 3, "criterio": "Si MAX(D,I,C) = 3"},
]

ESCALA_FRECUENCIA = [
    {"nivel": "Nula", "valor": 0.1, "descripcion": "Cada varios años"},
    {"nivel": "Baja", "valor": 1, "descripcion": "1 vez al año"},
    {"nivel": "Media", "valor": 2, "descripcion": "Mensualmente"},
    {"nivel": "Alta", "valor": 3, "descripcion": "A diario"},
]

ESCALA_DEGRADACION = [
    {"rango": "0.0", "descripcion": "Sin degradación"},
    {"rango": "0.1 - 0.3", "descripcion": "Degradación baja"},
    {"rango": "0.4 - 0.6", "descripcion": "Degradación media"},
    {"rango": "0.7 - 0.9", "descripcion": "Degradación alta"},
    {"rango": "1.0", "descripcion": "Degradación total"},
]

# Valor letra → número
VALOR_DIC = {"N": 0, "B": 1, "M": 2, "A": 3}
VALOR_DIC_INVERSO = {0: "N", 1: "B", 2: "M", 3: "A"}
VALOR_FREQ = {"N": 0.1, "B": 1, "M": 2, "A": 3}

# Límite de riesgo organizacional (constante)
LIMITE_RIESGO = 7.0

# Factor de reducción objetivo (50%)
FACTOR_REDUCCION = 0.5


# ==================== INICIALIZACIÓN ====================

def init_matriz_tables():
    """Inicializa las tablas específicas para el modelo de matriz"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # Tabla para criterios (referencia)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS CRITERIOS_VALORACION (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                Tipo_Criterio TEXT NOT NULL,
                Nivel TEXT NOT NULL,
                Valor REAL NOT NULL,
                Descripcion TEXT,
                Criterio TEXT
            )
        ''')
        
        # Tabla de valoración D/I/C por activo
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS IDENTIFICACION_VALORACION (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_Evaluacion TEXT NOT NULL,
                ID_Activo TEXT NOT NULL,
                Nombre_Activo TEXT,
                Dueno TEXT,
                Valor_Monetario REAL DEFAULT 0,
                Usuarios_Afectados INTEGER DEFAULT 0,
                D TEXT DEFAULT 'N',
                Valor_D INTEGER DEFAULT 0,
                I TEXT DEFAULT 'N',
                Valor_I INTEGER DEFAULT 0,
                C TEXT DEFAULT 'N',
                Valor_C INTEGER DEFAULT 0,
                Criticidad INTEGER DEFAULT 0,
                Criticidad_Nivel TEXT DEFAULT 'Nula',
                Fecha_Valoracion TEXT,
                FOREIGN KEY (ID_Evaluacion) REFERENCES EVALUACIONES(ID_Evaluacion),
                FOREIGN KEY (ID_Activo) REFERENCES INVENTARIO_ACTIVOS(ID_Activo),
                UNIQUE(ID_Evaluacion, ID_Activo)
            )
        ''')
        
        # Tabla de vulnerabilidades y amenazas con degradación
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS VULNERABILIDADES_AMENAZAS (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_Evaluacion TEXT NOT NULL,
                ID_Activo TEXT NOT NULL,
                Nombre_Activo TEXT,
                Criticidad INTEGER DEFAULT 0,
                Vulnerabilidad TEXT NOT NULL,
                Amenaza TEXT NOT NULL,
                Cod_Amenaza TEXT,
                Degradacion_D REAL DEFAULT 0,
                Degradacion_I REAL DEFAULT 0,
                Degradacion_C REAL DEFAULT 0,
                Impacto REAL DEFAULT 0,
                Fecha_Registro TEXT,
                FOREIGN KEY (ID_Evaluacion) REFERENCES EVALUACIONES(ID_Evaluacion),
                FOREIGN KEY (ID_Activo) REFERENCES INVENTARIO_ACTIVOS(ID_Activo)
            )
        ''')
        
        # Tabla de riesgo por amenaza
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS RIESGO_AMENAZA (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_Evaluacion TEXT NOT NULL,
                ID_Activo TEXT NOT NULL,
                Nombre_Activo TEXT,
                ID_Vulnerabilidad_Amenaza INTEGER,
                Amenaza TEXT NOT NULL,
                Frecuencia REAL DEFAULT 0.1,
                Frecuencia_Nivel TEXT DEFAULT 'Nula',
                Impacto REAL DEFAULT 0,
                Riesgo REAL DEFAULT 0,
                Fecha_Calculo TEXT,
                FOREIGN KEY (ID_Evaluacion) REFERENCES EVALUACIONES(ID_Evaluacion),
                FOREIGN KEY (ID_Activo) REFERENCES INVENTARIO_ACTIVOS(ID_Activo),
                FOREIGN KEY (ID_Vulnerabilidad_Amenaza) REFERENCES VULNERABILIDADES_AMENAZAS(id)
            )
        ''')
        
        # Tabla de mapa de riesgos (para visualización)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS MAPA_RIESGOS (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_Evaluacion TEXT NOT NULL,
                Riesgo_ID TEXT NOT NULL,
                ID_Activo TEXT NOT NULL,
                Nombre_Activo TEXT,
                Impacto REAL DEFAULT 0,
                Frecuencia REAL DEFAULT 0,
                Descripcion_Amenaza TEXT,
                Zona_Riesgo TEXT,
                Fecha_Registro TEXT,
                FOREIGN KEY (ID_Evaluacion) REFERENCES EVALUACIONES(ID_Evaluacion)
            )
        ''')
        
        # Tabla de riesgo agregado por activo
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS RIESGO_ACTIVOS (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_Evaluacion TEXT NOT NULL,
                ID_Activo TEXT NOT NULL,
                Nombre_Activo TEXT,
                Riesgo_Actual REAL DEFAULT 0,
                Riesgo_Objetivo REAL DEFAULT 0,
                Limite REAL DEFAULT 7,
                Estado TEXT DEFAULT 'Aceptable',
                Num_Amenazas INTEGER DEFAULT 0,
                Observacion TEXT,
                Fecha_Calculo TEXT,
                FOREIGN KEY (ID_Evaluacion) REFERENCES EVALUACIONES(ID_Evaluacion),
                FOREIGN KEY (ID_Activo) REFERENCES INVENTARIO_ACTIVOS(ID_Activo),
                UNIQUE(ID_Evaluacion, ID_Activo)
            )
        ''')
        
        # Tabla de salvaguardas recomendadas
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS SALVAGUARDAS (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_Evaluacion TEXT NOT NULL,
                ID_Activo TEXT NOT NULL,
                Nombre_Activo TEXT,
                Riesgo_ID TEXT,
                Vulnerabilidad TEXT,
                Amenaza TEXT,
                Salvaguarda TEXT NOT NULL,
                Prioridad TEXT DEFAULT 'Media',
                Estado TEXT DEFAULT 'Pendiente',
                Responsable TEXT,
                Fecha_Limite TEXT,
                Fecha_Registro TEXT,
                FOREIGN KEY (ID_Evaluacion) REFERENCES EVALUACIONES(ID_Evaluacion),
                FOREIGN KEY (ID_Activo) REFERENCES INVENTARIO_ACTIVOS(ID_Activo)
            )
        ''')
        
        conn.commit()
        
    # Poblar criterios de referencia si no existen
    _poblar_criterios_referencia()


def _poblar_criterios_referencia():
    """Pobla la tabla de criterios de valoración con las escalas de referencia"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # Verificar si ya existen criterios
        cursor.execute("SELECT COUNT(*) FROM CRITERIOS_VALORACION")
        if cursor.fetchone()[0] > 0:
            return
        
        # Insertar escalas de Disponibilidad
        for item in ESCALA_DISPONIBILIDAD:
            cursor.execute('''
                INSERT INTO CRITERIOS_VALORACION (Tipo_Criterio, Nivel, Valor, Descripcion)
                VALUES (?, ?, ?, ?)
            ''', ("Disponibilidad", item["nivel"], item["valor"], item["descripcion"]))
        
        # Insertar escalas de Integridad
        for item in ESCALA_INTEGRIDAD:
            cursor.execute('''
                INSERT INTO CRITERIOS_VALORACION (Tipo_Criterio, Nivel, Valor, Descripcion)
                VALUES (?, ?, ?, ?)
            ''', ("Integridad", item["nivel"], item["valor"], item["descripcion"]))
        
        # Insertar escalas de Confidencialidad
        for item in ESCALA_CONFIDENCIALIDAD:
            cursor.execute('''
                INSERT INTO CRITERIOS_VALORACION (Tipo_Criterio, Nivel, Valor, Descripcion)
                VALUES (?, ?, ?, ?)
            ''', ("Confidencialidad", item["nivel"], item["valor"], item["descripcion"]))
        
        # Insertar escalas de Criticidad
        for item in ESCALA_CRITICIDAD:
            cursor.execute('''
                INSERT INTO CRITERIOS_VALORACION (Tipo_Criterio, Nivel, Valor, Descripcion, Criterio)
                VALUES (?, ?, ?, ?, ?)
            ''', ("Criticidad", item["nivel"], item["valor"], None, item["criterio"]))
        
        # Insertar escalas de Frecuencia
        for item in ESCALA_FRECUENCIA:
            cursor.execute('''
                INSERT INTO CRITERIOS_VALORACION (Tipo_Criterio, Nivel, Valor, Descripcion)
                VALUES (?, ?, ?, ?)
            ''', ("Frecuencia", item["nivel"], item["valor"], item["descripcion"]))
        
        conn.commit()


# ==================== CRITERIOS DE VALORACIÓN ====================

def get_criterios_valoracion() -> Dict[str, pd.DataFrame]:
    """Obtiene todos los criterios de valoración agrupados por tipo"""
    with get_connection() as conn:
        df = pd.read_sql_query("SELECT * FROM CRITERIOS_VALORACION ORDER BY Tipo_Criterio, Valor", conn)
    
    result = {}
    for tipo in ["Disponibilidad", "Integridad", "Confidencialidad", "Criticidad", "Frecuencia"]:
        result[tipo] = df[df["Tipo_Criterio"] == tipo].reset_index(drop=True)
    
    return result


def get_escala_degradacion() -> List[Dict]:
    """Retorna la escala de degradación (no se almacena en DB)"""
    return ESCALA_DEGRADACION


# ==================== ACTIVOS (FORMATO MATRIZ) ====================

def get_activos_matriz(id_evaluacion: str) -> pd.DataFrame:
    """Obtiene activos en formato de la matriz de referencia"""
    query = """
        SELECT 
            a.ID_Activo,
            a.ID_Evaluacion,
            a.Nombre_Activo,
            a.Tipo_Activo,
            a.Ubicacion,
            a.Propietario as Area_Responsable,
            a.Tipo_Servicio as Finalidad_Uso,
            a.App_Critica,
            a.Estado,
            a.Fecha_Creacion as Fecha_Instalacion,
            COALESCE(v.Criticidad, 0) as Criticidad,
            COALESCE(v.Criticidad_Nivel, 'Pendiente') as Criticidad_Nivel
        FROM INVENTARIO_ACTIVOS a
        LEFT JOIN IDENTIFICACION_VALORACION v 
            ON a.ID_Activo = v.ID_Activo AND a.ID_Evaluacion = v.ID_Evaluacion
        WHERE a.ID_Evaluacion = ?
        ORDER BY a.Nombre_Activo
    """
    with get_connection() as conn:
        df = pd.read_sql_query(query, conn, params=(id_evaluacion,))
    return df


# ==================== IDENTIFICACIÓN Y VALORACIÓN ====================

def guardar_valoracion_dic(
    id_evaluacion: str,
    id_activo: str,
    nombre_activo: str,
    dueno: str = "",
    valor_monetario: float = 0,
    usuarios_afectados: int = 0,
    d_nivel: str = "N",
    i_nivel: str = "N",
    c_nivel: str = "N"
) -> bool:
    """Guarda o actualiza la valoración D/I/C de un activo"""
    valor_d = VALOR_DIC.get(d_nivel.upper(), 0)
    valor_i = VALOR_DIC.get(i_nivel.upper(), 0)
    valor_c = VALOR_DIC.get(c_nivel.upper(), 0)
    criticidad = max(valor_d, valor_i, valor_c)
    criticidad_nivel = ESCALA_CRITICIDAD[criticidad]["nivel"]
    
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # Verificar si ya existe
        cursor.execute('''
            SELECT id FROM IDENTIFICACION_VALORACION 
            WHERE ID_Evaluacion = ? AND ID_Activo = ?
        ''', (id_evaluacion, id_activo))
        
        existe = cursor.fetchone()
        
        if existe:
            cursor.execute('''
                UPDATE IDENTIFICACION_VALORACION SET
                    Nombre_Activo = ?,
                    Dueno = ?,
                    Valor_Monetario = ?,
                    Usuarios_Afectados = ?,
                    D = ?, Valor_D = ?,
                    I = ?, Valor_I = ?,
                    C = ?, Valor_C = ?,
                    Criticidad = ?,
                    Criticidad_Nivel = ?,
                    Fecha_Valoracion = ?
                WHERE ID_Evaluacion = ? AND ID_Activo = ?
            ''', (
                nombre_activo, dueno, valor_monetario, usuarios_afectados,
                d_nivel.upper(), valor_d,
                i_nivel.upper(), valor_i,
                c_nivel.upper(), valor_c,
                criticidad, criticidad_nivel,
                dt.datetime.now().isoformat(),
                id_evaluacion, id_activo
            ))
        else:
            cursor.execute('''
                INSERT INTO IDENTIFICACION_VALORACION 
                (ID_Evaluacion, ID_Activo, Nombre_Activo, Dueno, Valor_Monetario,
                 Usuarios_Afectados, D, Valor_D, I, Valor_I, C, Valor_C,
                 Criticidad, Criticidad_Nivel, Fecha_Valoracion)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                id_evaluacion, id_activo, nombre_activo, dueno, valor_monetario,
                usuarios_afectados, d_nivel.upper(), valor_d, i_nivel.upper(), valor_i,
                c_nivel.upper(), valor_c, criticidad, criticidad_nivel,
                dt.datetime.now().isoformat()
            ))
        
        conn.commit()
    
    return True


def get_valoraciones_evaluacion(id_evaluacion: str) -> pd.DataFrame:
    """Obtiene todas las valoraciones D/I/C de una evaluación"""
    query = """
        SELECT 
            v.*,
            a.Tipo_Activo,
            a.Ubicacion
        FROM IDENTIFICACION_VALORACION v
        JOIN INVENTARIO_ACTIVOS a ON v.ID_Activo = a.ID_Activo
        WHERE v.ID_Evaluacion = ?
        ORDER BY v.Criticidad DESC, v.Nombre_Activo
    """
    with get_connection() as conn:
        df = pd.read_sql_query(query, conn, params=(id_evaluacion,))
    return df


def get_valoracion_activo(id_evaluacion: str, id_activo: str) -> Optional[Dict]:
    """Obtiene la valoración de un activo específico"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT * FROM IDENTIFICACION_VALORACION 
            WHERE ID_Evaluacion = ? AND ID_Activo = ?
        ''', (id_evaluacion, id_activo))
        row = cursor.fetchone()
        if row:
            return dict(row)
    return None


# ==================== VULNERABILIDADES Y AMENAZAS ====================

def agregar_vulnerabilidad_amenaza(
    id_evaluacion: str,
    id_activo: str,
    nombre_activo: str,
    vulnerabilidad: str,
    amenaza: str,
    cod_amenaza: str = "",
    deg_d: float = 0.0,
    deg_i: float = 0.0,
    deg_c: float = 0.0
) -> int:
    """Agrega una vulnerabilidad-amenaza con su degradación"""
    # Obtener criticidad del activo
    valoracion = get_valoracion_activo(id_evaluacion, id_activo)
    criticidad = valoracion["Criticidad"] if valoracion else 0
    
    # Calcular impacto: CRITICIDAD × MAX(Deg_D, Deg_I, Deg_C)
    impacto = criticidad * max(deg_d, deg_i, deg_c)
    
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO VULNERABILIDADES_AMENAZAS 
            (ID_Evaluacion, ID_Activo, Nombre_Activo, Criticidad, 
             Vulnerabilidad, Amenaza, Cod_Amenaza, 
             Degradacion_D, Degradacion_I, Degradacion_C, Impacto, Fecha_Registro)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            id_evaluacion, id_activo, nombre_activo, criticidad,
            vulnerabilidad, amenaza, cod_amenaza,
            deg_d, deg_i, deg_c, impacto,
            dt.datetime.now().isoformat()
        ))
        conn.commit()
        return cursor.lastrowid


def actualizar_vulnerabilidad_amenaza(
    id_va: int,
    vulnerabilidad: str = None,
    amenaza: str = None,
    deg_d: float = None,
    deg_i: float = None,
    deg_c: float = None
) -> bool:
    """Actualiza una vulnerabilidad-amenaza existente"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # Obtener datos actuales
        cursor.execute("SELECT * FROM VULNERABILIDADES_AMENAZAS WHERE id = ?", (id_va,))
        row = cursor.fetchone()
        if not row:
            return False
        
        current = dict(row)
        vulnerabilidad = vulnerabilidad if vulnerabilidad is not None else current["Vulnerabilidad"]
        amenaza = amenaza if amenaza is not None else current["Amenaza"]
        deg_d = deg_d if deg_d is not None else current["Degradacion_D"]
        deg_i = deg_i if deg_i is not None else current["Degradacion_I"]
        deg_c = deg_c if deg_c is not None else current["Degradacion_C"]
        
        # Recalcular impacto
        impacto = current["Criticidad"] * max(deg_d, deg_i, deg_c)
        
        cursor.execute('''
            UPDATE VULNERABILIDADES_AMENAZAS SET
                Vulnerabilidad = ?,
                Amenaza = ?,
                Degradacion_D = ?,
                Degradacion_I = ?,
                Degradacion_C = ?,
                Impacto = ?
            WHERE id = ?
        ''', (vulnerabilidad, amenaza, deg_d, deg_i, deg_c, impacto, id_va))
        
        conn.commit()
    return True


def eliminar_vulnerabilidad_amenaza(id_va: int) -> bool:
    """Elimina una vulnerabilidad-amenaza"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM VULNERABILIDADES_AMENAZAS WHERE id = ?", (id_va,))
        conn.commit()
        return cursor.rowcount > 0


def get_vulnerabilidades_activo(id_evaluacion: str, id_activo: str) -> pd.DataFrame:
    """Obtiene vulnerabilidades y amenazas de un activo"""
    query = """
        SELECT * FROM VULNERABILIDADES_AMENAZAS 
        WHERE ID_Evaluacion = ? AND ID_Activo = ?
        ORDER BY Impacto DESC
    """
    with get_connection() as conn:
        df = pd.read_sql_query(query, conn, params=(id_evaluacion, id_activo))
    return df


def get_vulnerabilidades_evaluacion(id_evaluacion: str) -> pd.DataFrame:
    """Obtiene todas las vulnerabilidades y amenazas de una evaluación"""
    query = """
        SELECT * FROM VULNERABILIDADES_AMENAZAS 
        WHERE ID_Evaluacion = ?
        ORDER BY Nombre_Activo, Impacto DESC
    """
    with get_connection() as conn:
        df = pd.read_sql_query(query, conn, params=(id_evaluacion,))
    return df


# ==================== RIESGO POR AMENAZA ====================

def calcular_riesgo_amenaza(
    id_evaluacion: str,
    id_activo: str,
    id_va: int,
    frecuencia: float = 0.1
) -> float:
    """Calcula el riesgo para una amenaza específica: Frecuencia × Impacto"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # Obtener la vulnerabilidad-amenaza
        cursor.execute("SELECT * FROM VULNERABILIDADES_AMENAZAS WHERE id = ?", (id_va,))
        va = cursor.fetchone()
        if not va:
            return 0.0
        
        va_dict = dict(va)
        impacto = va_dict["Impacto"]
        riesgo = frecuencia * impacto
        
        # Determinar nivel de frecuencia
        freq_nivel = "Nula"
        for item in reversed(ESCALA_FRECUENCIA):
            if frecuencia >= item["valor"]:
                freq_nivel = item["nivel"]
                break
        
        # Guardar o actualizar en RIESGO_AMENAZA
        cursor.execute('''
            SELECT id FROM RIESGO_AMENAZA WHERE ID_Vulnerabilidad_Amenaza = ?
        ''', (id_va,))
        existe = cursor.fetchone()
        
        if existe:
            cursor.execute('''
                UPDATE RIESGO_AMENAZA SET
                    Frecuencia = ?,
                    Frecuencia_Nivel = ?,
                    Impacto = ?,
                    Riesgo = ?,
                    Fecha_Calculo = ?
                WHERE ID_Vulnerabilidad_Amenaza = ?
            ''', (frecuencia, freq_nivel, impacto, riesgo, dt.datetime.now().isoformat(), id_va))
        else:
            cursor.execute('''
                INSERT INTO RIESGO_AMENAZA 
                (ID_Evaluacion, ID_Activo, Nombre_Activo, ID_Vulnerabilidad_Amenaza,
                 Amenaza, Frecuencia, Frecuencia_Nivel, Impacto, Riesgo, Fecha_Calculo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                id_evaluacion, id_activo, va_dict["Nombre_Activo"], id_va,
                va_dict["Amenaza"], frecuencia, freq_nivel, impacto, riesgo,
                dt.datetime.now().isoformat()
            ))
        
        conn.commit()
        return riesgo


def get_riesgos_activo(id_evaluacion: str, id_activo: str) -> pd.DataFrame:
    """Obtiene todos los riesgos calculados de un activo"""
    query = """
        SELECT 
            r.*,
            va.Vulnerabilidad,
            va.Degradacion_D,
            va.Degradacion_I,
            va.Degradacion_C
        FROM RIESGO_AMENAZA r
        JOIN VULNERABILIDADES_AMENAZAS va ON r.ID_Vulnerabilidad_Amenaza = va.id
        WHERE r.ID_Evaluacion = ? AND r.ID_Activo = ?
        ORDER BY r.Riesgo DESC
    """
    with get_connection() as conn:
        df = pd.read_sql_query(query, conn, params=(id_evaluacion, id_activo))
    return df


def get_riesgos_evaluacion(id_evaluacion: str) -> pd.DataFrame:
    """Obtiene todos los riesgos de una evaluación"""
    query = """
        SELECT 
            r.*,
            va.Vulnerabilidad,
            va.Cod_Amenaza,
            va.Degradacion_D,
            va.Degradacion_I,
            va.Degradacion_C,
            va.Criticidad
        FROM RIESGO_AMENAZA r
        JOIN VULNERABILIDADES_AMENAZAS va ON r.ID_Vulnerabilidad_Amenaza = va.id
        WHERE r.ID_Evaluacion = ?
        ORDER BY r.Riesgo DESC
    """
    with get_connection() as conn:
        df = pd.read_sql_query(query, conn, params=(id_evaluacion,))
    return df


# ==================== MAPA DE RIESGOS ====================

def generar_mapa_riesgos(id_evaluacion: str) -> pd.DataFrame:
    """Genera el mapa de riesgos (Impacto vs Frecuencia) para visualización"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # Obtener todos los riesgos
        riesgos = get_riesgos_evaluacion(id_evaluacion)
        
        if riesgos.empty:
            return pd.DataFrame()
        
        # Limpiar mapa anterior
        cursor.execute("DELETE FROM MAPA_RIESGOS WHERE ID_Evaluacion = ?", (id_evaluacion,))
        
        # Generar nuevo mapa
        for idx, row in riesgos.iterrows():
            riesgo_id = f"R{idx + 1}"
            
            # Determinar zona de riesgo
            riesgo_val = row["Riesgo"]
            if riesgo_val >= 6:
                zona = "Crítico"
            elif riesgo_val >= 4:
                zona = "Alto"
            elif riesgo_val >= 2:
                zona = "Medio"
            else:
                zona = "Bajo"
            
            cursor.execute('''
                INSERT INTO MAPA_RIESGOS 
                (ID_Evaluacion, Riesgo_ID, ID_Activo, Nombre_Activo, 
                 Impacto, Frecuencia, Descripcion_Amenaza, Zona_Riesgo, Fecha_Registro)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                id_evaluacion, riesgo_id, row["ID_Activo"], row["Nombre_Activo"],
                row["Impacto"], row["Frecuencia"], row["Amenaza"], zona,
                dt.datetime.now().isoformat()
            ))
        
        conn.commit()
    
    # Retornar el mapa generado
    return get_mapa_riesgos(id_evaluacion)


def get_mapa_riesgos(id_evaluacion: str) -> pd.DataFrame:
    """Obtiene el mapa de riesgos de una evaluación"""
    query = """
        SELECT * FROM MAPA_RIESGOS 
        WHERE ID_Evaluacion = ?
        ORDER BY Impacto DESC, Frecuencia DESC
    """
    with get_connection() as conn:
        df = pd.read_sql_query(query, conn, params=(id_evaluacion,))
    return df


# ==================== RIESGO AGREGADO POR ACTIVO ====================

def calcular_riesgo_activo(id_evaluacion: str, id_activo: str) -> Dict:
    """
    Calcula el riesgo agregado de un activo:
    - Riesgo_Actual = PROMEDIO(todos los riesgos del activo)
    - Riesgo_Objetivo = Riesgo_Actual × 0.5
    - Límite = 7 (constante organizacional)
    """
    riesgos = get_riesgos_activo(id_evaluacion, id_activo)
    
    if riesgos.empty:
        return {
            "riesgo_actual": 0,
            "riesgo_objetivo": 0,
            "limite": LIMITE_RIESGO,
            "estado": "Sin evaluar",
            "num_amenazas": 0
        }
    
    riesgo_actual = riesgos["Riesgo"].mean()
    riesgo_objetivo = riesgo_actual * FACTOR_REDUCCION
    
    if riesgo_actual > LIMITE_RIESGO:
        estado = "Tratamiento Urgente"
    elif riesgo_actual > LIMITE_RIESGO * 0.7:
        estado = "Atención Requerida"
    else:
        estado = "Aceptable"
    
    # Guardar en BD
    with get_connection() as conn:
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id FROM RIESGO_ACTIVOS WHERE ID_Evaluacion = ? AND ID_Activo = ?
        ''', (id_evaluacion, id_activo))
        existe = cursor.fetchone()
        
        # Obtener nombre del activo
        cursor.execute("SELECT Nombre_Activo FROM INVENTARIO_ACTIVOS WHERE ID_Activo = ?", (id_activo,))
        activo_row = cursor.fetchone()
        nombre_activo = activo_row[0] if activo_row else ""
        
        if existe:
            cursor.execute('''
                UPDATE RIESGO_ACTIVOS SET
                    Nombre_Activo = ?,
                    Riesgo_Actual = ?,
                    Riesgo_Objetivo = ?,
                    Limite = ?,
                    Estado = ?,
                    Num_Amenazas = ?,
                    Fecha_Calculo = ?
                WHERE ID_Evaluacion = ? AND ID_Activo = ?
            ''', (
                nombre_activo, riesgo_actual, riesgo_objetivo, LIMITE_RIESGO,
                estado, len(riesgos), dt.datetime.now().isoformat(),
                id_evaluacion, id_activo
            ))
        else:
            cursor.execute('''
                INSERT INTO RIESGO_ACTIVOS 
                (ID_Evaluacion, ID_Activo, Nombre_Activo, Riesgo_Actual,
                 Riesgo_Objetivo, Limite, Estado, Num_Amenazas, Fecha_Calculo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                id_evaluacion, id_activo, nombre_activo, riesgo_actual,
                riesgo_objetivo, LIMITE_RIESGO, estado, len(riesgos),
                dt.datetime.now().isoformat()
            ))
        
        conn.commit()
    
    return {
        "riesgo_actual": riesgo_actual,
        "riesgo_objetivo": riesgo_objetivo,
        "limite": LIMITE_RIESGO,
        "estado": estado,
        "num_amenazas": len(riesgos)
    }


def get_riesgos_activos_evaluacion(id_evaluacion: str) -> pd.DataFrame:
    """Obtiene el riesgo agregado de todos los activos de una evaluación"""
    query = """
        SELECT * FROM RIESGO_ACTIVOS 
        WHERE ID_Evaluacion = ?
        ORDER BY Riesgo_Actual DESC
    """
    with get_connection() as conn:
        df = pd.read_sql_query(query, conn, params=(id_evaluacion,))
    return df


def recalcular_todos_riesgos_activos(id_evaluacion: str) -> int:
    """Recalcula el riesgo de todos los activos de una evaluación"""
    # Obtener todos los activos con valoración
    valoraciones = get_valoraciones_evaluacion(id_evaluacion)
    
    count = 0
    for _, val in valoraciones.iterrows():
        calcular_riesgo_activo(id_evaluacion, val["ID_Activo"])
        count += 1
    
    return count


# ==================== SALVAGUARDAS ====================

def agregar_salvaguarda(
    id_evaluacion: str,
    id_activo: str,
    nombre_activo: str,
    salvaguarda: str,
    riesgo_id: str = "",
    vulnerabilidad: str = "",
    amenaza: str = "",
    prioridad: str = "Media",
    responsable: str = "",
    fecha_limite: str = ""
) -> int:
    """Agrega una salvaguarda/control recomendado"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO SALVAGUARDAS 
            (ID_Evaluacion, ID_Activo, Nombre_Activo, Riesgo_ID, 
             Vulnerabilidad, Amenaza, Salvaguarda, Prioridad, 
             Estado, Responsable, Fecha_Limite, Fecha_Registro)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            id_evaluacion, id_activo, nombre_activo, riesgo_id,
            vulnerabilidad, amenaza, salvaguarda, prioridad,
            "Pendiente", responsable, fecha_limite,
            dt.datetime.now().isoformat()
        ))
        conn.commit()
        return cursor.lastrowid


def actualizar_estado_salvaguarda(id_salvaguarda: int, estado: str) -> bool:
    """Actualiza el estado de una salvaguarda"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE SALVAGUARDAS SET Estado = ? WHERE id = ?
        ''', (estado, id_salvaguarda))
        conn.commit()
        return cursor.rowcount > 0


def eliminar_salvaguarda(id_salvaguarda: int) -> bool:
    """Elimina una salvaguarda"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM SALVAGUARDAS WHERE id = ?", (id_salvaguarda,))
        conn.commit()
        return cursor.rowcount > 0


def get_salvaguardas_activo(id_evaluacion: str, id_activo: str) -> pd.DataFrame:
    """Obtiene las salvaguardas de un activo"""
    query = """
        SELECT * FROM SALVAGUARDAS 
        WHERE ID_Evaluacion = ? AND ID_Activo = ?
        ORDER BY 
            CASE Prioridad 
                WHEN 'Alta' THEN 1 
                WHEN 'Media' THEN 2 
                WHEN 'Baja' THEN 3 
            END,
            Estado
    """
    with get_connection() as conn:
        df = pd.read_sql_query(query, conn, params=(id_evaluacion, id_activo))
    return df


def get_salvaguardas_evaluacion(id_evaluacion: str) -> pd.DataFrame:
    """Obtiene todas las salvaguardas de una evaluación"""
    query = """
        SELECT * FROM SALVAGUARDAS 
        WHERE ID_Evaluacion = ?
        ORDER BY Nombre_Activo, 
            CASE Prioridad 
                WHEN 'Alta' THEN 1 
                WHEN 'Media' THEN 2 
                WHEN 'Baja' THEN 3 
            END
    """
    with get_connection() as conn:
        df = pd.read_sql_query(query, conn, params=(id_evaluacion,))
    return df


# ==================== ESTADÍSTICAS ====================

def get_estadisticas_evaluacion_matriz(id_evaluacion: str) -> Dict:
    """Obtiene estadísticas generales de una evaluación en el modelo matriz"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # Activos totales
        cursor.execute('''
            SELECT COUNT(*) FROM INVENTARIO_ACTIVOS WHERE ID_Evaluacion = ?
        ''', (id_evaluacion,))
        total_activos = cursor.fetchone()[0]
        
        # Activos valorados
        cursor.execute('''
            SELECT COUNT(*) FROM IDENTIFICACION_VALORACION WHERE ID_Evaluacion = ?
        ''', (id_evaluacion,))
        activos_valorados = cursor.fetchone()[0]
        
        # Vulnerabilidades/Amenazas
        cursor.execute('''
            SELECT COUNT(*) FROM VULNERABILIDADES_AMENAZAS WHERE ID_Evaluacion = ?
        ''', (id_evaluacion,))
        total_vulnerabilidades = cursor.fetchone()[0]
        
        # Riesgos calculados
        cursor.execute('''
            SELECT COUNT(*) FROM RIESGO_AMENAZA WHERE ID_Evaluacion = ?
        ''', (id_evaluacion,))
        riesgos_calculados = cursor.fetchone()[0]
        
        # Salvaguardas
        cursor.execute('''
            SELECT COUNT(*) FROM SALVAGUARDAS WHERE ID_Evaluacion = ?
        ''', (id_evaluacion,))
        total_salvaguardas = cursor.fetchone()[0]
        
        # Salvaguardas implementadas
        cursor.execute('''
            SELECT COUNT(*) FROM SALVAGUARDAS 
            WHERE ID_Evaluacion = ? AND Estado = 'Implementada'
        ''', (id_evaluacion,))
        salvaguardas_impl = cursor.fetchone()[0]
        
        # Activos en estado urgente
        cursor.execute('''
            SELECT COUNT(*) FROM RIESGO_ACTIVOS 
            WHERE ID_Evaluacion = ? AND Estado = 'Tratamiento Urgente'
        ''', (id_evaluacion,))
        activos_urgentes = cursor.fetchone()[0]
        
        # Riesgo promedio
        cursor.execute('''
            SELECT AVG(Riesgo_Actual) FROM RIESGO_ACTIVOS WHERE ID_Evaluacion = ?
        ''', (id_evaluacion,))
        riesgo_promedio = cursor.fetchone()[0] or 0
    
    return {
        "total_activos": total_activos,
        "activos_valorados": activos_valorados,
        "pct_valorados": (activos_valorados / total_activos * 100) if total_activos > 0 else 0,
        "total_vulnerabilidades": total_vulnerabilidades,
        "riesgos_calculados": riesgos_calculados,
        "total_salvaguardas": total_salvaguardas,
        "salvaguardas_implementadas": salvaguardas_impl,
        "pct_salvaguardas": (salvaguardas_impl / total_salvaguardas * 100) if total_salvaguardas > 0 else 0,
        "activos_urgentes": activos_urgentes,
        "riesgo_promedio": riesgo_promedio,
        "limite_riesgo": LIMITE_RIESGO
    }


# ==================== EXPORTACIÓN ====================

def exportar_matriz_excel(id_evaluacion: str, nombre_evaluacion: str = "Evaluacion") -> bytes:
    """Exporta la matriz completa a Excel con múltiples hojas"""
    import io
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    
    # Hoja 1: Criterios de Valoración
    ws1 = wb.active
    ws1.title = "CRITERIOS_VALORACION"
    criterios = get_criterios_valoracion()
    row_num = 1
    for tipo, df in criterios.items():
        ws1.cell(row=row_num, column=1, value=tipo).font = Font(bold=True)
        row_num += 1
        for r in dataframe_to_rows(df[["Nivel", "Valor", "Descripcion"]], index=False, header=True):
            for c_idx, value in enumerate(r, 1):
                ws1.cell(row=row_num, column=c_idx, value=value)
            row_num += 1
        row_num += 1
    
    # Hoja 2: Activos
    ws2 = wb.create_sheet("ACTIVOS")
    activos = get_activos_matriz(id_evaluacion)
    for r in dataframe_to_rows(activos, index=False, header=True):
        ws2.append(r)
    
    # Hoja 3: Identificación y Valoración
    ws3 = wb.create_sheet("IDENTIFICACION_VALORACION")
    valoraciones = get_valoraciones_evaluacion(id_evaluacion)
    for r in dataframe_to_rows(valoraciones, index=False, header=True):
        ws3.append(r)
    
    # Hoja 4: Vulnerabilidades y Amenazas
    ws4 = wb.create_sheet("VULNERABILIDADES_AMENAZAS")
    vulns = get_vulnerabilidades_evaluacion(id_evaluacion)
    for r in dataframe_to_rows(vulns, index=False, header=True):
        ws4.append(r)
    
    # Hoja 5: Riesgo
    ws5 = wb.create_sheet("RIESGO")
    riesgos = get_riesgos_evaluacion(id_evaluacion)
    for r in dataframe_to_rows(riesgos, index=False, header=True):
        ws5.append(r)
    
    # Hoja 6: Mapa de Riesgos
    ws6 = wb.create_sheet("MAPA_RIESGOS")
    mapa = get_mapa_riesgos(id_evaluacion)
    for r in dataframe_to_rows(mapa, index=False, header=True):
        ws6.append(r)
    
    # Hoja 7: Riesgo por Activos
    ws7 = wb.create_sheet("RIESGO_ACTIVOS")
    riesgos_act = get_riesgos_activos_evaluacion(id_evaluacion)
    for r in dataframe_to_rows(riesgos_act, index=False, header=True):
        ws7.append(r)
    
    # Hoja 8: Salvaguardas
    ws8 = wb.create_sheet("SALVAGUARDAS")
    salvs = get_salvaguardas_evaluacion(id_evaluacion)
    for r in dataframe_to_rows(salvs, index=False, header=True):
        ws8.append(r)
    
    # Guardar a bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
