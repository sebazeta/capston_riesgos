"""
Servicio de gestión de archivos Excel con escritura atómica y sistema de bloqueo
"""
import datetime as dt
import pandas as pd
import shutil
import os
import tempfile
from openpyxl import load_workbook, Workbook
from typing import List, Dict, Any
from contextlib import contextmanager
import threading

EXCEL_PATH = "matriz_riesgos_v2.xlsx"
BACKUP_PATH = "matriz_riesgos_v2_backup.xlsx"

# Lock global para evitar escrituras concurrentes en el mismo proceso
_file_lock = threading.Lock()


@contextmanager
def file_lock():
    """Context manager para bloquear acceso al archivo Excel"""
    _file_lock.acquire()
    try:
        yield
    finally:
        _file_lock.release()


def crear_backup():
    """Crea una copia de respaldo del archivo Excel"""
    if os.path.exists(EXCEL_PATH):
        try:
            shutil.copy2(EXCEL_PATH, BACKUP_PATH)
        except Exception:
            pass


def restaurar_backup():
    """Restaura el archivo desde el backup si está corrupto"""
    if os.path.exists(BACKUP_PATH):
        try:
            shutil.copy2(BACKUP_PATH, EXCEL_PATH)
            return True
        except Exception:
            pass
    return False


def verificar_archivo_valido(path: str) -> bool:
    """Verifica si un archivo Excel es válido"""
    try:
        wb = load_workbook(path)
        wb.close()
        return True
    except Exception:
        return False


def guardar_workbook_atomico(wb: Workbook):
    """Guarda el workbook de forma atómica usando archivo temporal"""
    # Crear archivo temporal en el mismo directorio
    temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx', dir=os.path.dirname(EXCEL_PATH) or '.')
    os.close(temp_fd)
    
    try:
        # Guardar al archivo temporal
        wb.save(temp_path)
        wb.close()
        
        # Verificar que el archivo temporal es válido
        if not verificar_archivo_valido(temp_path):
            os.unlink(temp_path)
            raise Exception("El archivo temporal no es válido")
        
        # Renombrar atómicamente (en Windows necesitamos borrar primero)
        if os.path.exists(EXCEL_PATH):
            os.unlink(EXCEL_PATH)
        shutil.move(temp_path, EXCEL_PATH)
        
        # Crear backup después de escritura exitosa
        crear_backup()
        
    except Exception as e:
        # Limpiar archivo temporal si existe
        if os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except:
                pass
        raise e


def ensure_workbook():
    """Asegura que el workbook Excel existe y no está corrupto"""
    # Intentar abrir el archivo existente
    if os.path.exists(EXCEL_PATH):
        if verificar_archivo_valido(EXCEL_PATH):
            return load_workbook(EXCEL_PATH)
        else:
            # Archivo corrupto - intentar restaurar desde backup
            if restaurar_backup() and verificar_archivo_valido(EXCEL_PATH):
                return load_workbook(EXCEL_PATH)
            else:
                # Eliminar archivo corrupto
                try:
                    os.unlink(EXCEL_PATH)
                except:
                    pass
    
    # Crear archivo nuevo si no existe
    raise FileNotFoundError(
        "Archivo Excel no existe o está corrupto. "
        "Ejecuta: python init_proyecto.py"
    )


def ensure_sheet_exists(sheet_name: str, headers: list):
    """Crea una hoja si no existe"""
    with file_lock():
        wb = load_workbook(EXCEL_PATH)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
            guardar_workbook_atomico(wb)
        else:
            wb.close()


def read_sheet(sheet_name: str) -> pd.DataFrame:
    """Lee una hoja de Excel como DataFrame"""
    with file_lock():
        if not os.path.exists(EXCEL_PATH):
            return pd.DataFrame()
        
        if not verificar_archivo_valido(EXCEL_PATH):
            if restaurar_backup():
                pass
            else:
                return pd.DataFrame()
        
        try:
            wb = load_workbook(EXCEL_PATH)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return pd.DataFrame()
            wb.close()
            return pd.read_excel(EXCEL_PATH, sheet_name=sheet_name)
        except Exception as e:
            print(f"Error leyendo hoja {sheet_name}: {e}")
            return pd.DataFrame()


def append_rows(sheet_name: str, rows: List[Dict[str, Any]]):
    """Añade filas a una hoja Excel de forma atómica"""
    if not rows:
        return
    
    with file_lock():
        wb = load_workbook(EXCEL_PATH)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet {sheet_name} no existe. Crea con ensure_sheet_exists() primero.")
        
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        for r in rows:
            ws.append([r.get(h) for h in headers])
        
        guardar_workbook_atomico(wb)


def update_row(sheet_name: str, key_column: str, key_value: str, updates: Dict[str, Any]):
    """Actualiza una fila específica de forma atómica"""
    with file_lock():
        wb = load_workbook(EXCEL_PATH)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return False
        
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        key_idx = headers.index(key_column) + 1 if key_column in headers else None
        
        if key_idx is None:
            wb.close()
            return False
        
        found = False
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, key_idx).value) == str(key_value):
                for col_name, value in updates.items():
                    if col_name in headers:
                        col_idx = headers.index(col_name) + 1
                        ws.cell(row, col_idx).value = value
                found = True
                break
        
        if found:
            guardar_workbook_atomico(wb)
        else:
            wb.close()
        
        return found


def delete_row(sheet_name: str, key_column: str, key_value: str):
    """Elimina una fila específica de forma atómica"""
    with file_lock():
        wb = load_workbook(EXCEL_PATH)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return False
        
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        key_idx = headers.index(key_column) + 1 if key_column in headers else None
        
        if key_idx is None:
            wb.close()
            return False
        
        row_to_delete = None
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, key_idx).value) == str(key_value):
                row_to_delete = row
                break
        
        if row_to_delete:
            ws.delete_rows(row_to_delete)
            guardar_workbook_atomico(wb)
            return True
        
        wb.close()
        return False


def set_eval_active(eval_id: str, nombre: str):
    """Establece una evaluación como activa de forma atómica"""
    with file_lock():
        wb = load_workbook(EXCEL_PATH)
        
        if "EVALUACIONES" not in wb.sheetnames:
            ws = wb.create_sheet("EVALUACIONES")
            ws.append(["ID_Evaluacion", "Nombre", "Fecha", "Estado", "Descripcion"])
        else:
            ws = wb["EVALUACIONES"]
        
        headers = [c.value for c in ws[1]]
        idx = {h: i + 1 for i, h in enumerate(headers)}

        # set all inactive
        for r in range(2, ws.max_row + 1):
            ws.cell(r, idx.get("Estado", 4)).value = "Inactiva"

        found = None
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, idx["ID_Evaluacion"]).value) == str(eval_id):
                found = r
                break

        now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row = {
            "ID_Evaluacion": eval_id,
            "Nombre": nombre,
            "Fecha": now,
            "Estado": "Activa",
            "Descripcion": "Evaluación activa desde la GUI"
        }

        if found is None:
            ws.append([row.get(h) for h in headers])
        else:
            for h, v in row.items():
                ws.cell(found, idx[h]).value = v

        guardar_workbook_atomico(wb)


def update_cuestionarios_version(eval_id: str, activo_id: str, fecha: str):
    """Actualiza la versión del cuestionario"""
    with file_lock():
        wb = load_workbook(EXCEL_PATH)
        if "CUESTIONARIOS" not in wb.sheetnames:
            wb.close()
            return
        
        ws = wb["CUESTIONARIOS"]
        headers = [c.value for c in ws[1]]
        
        if "Fecha_Version" not in headers:
            wb.close()
            return
        
        idx_eval = headers.index("ID_Evaluacion") + 1
        idx_activo = headers.index("ID_Activo") + 1
        idx_fecha = headers.index("Fecha_Version") + 1
        
        for row in range(2, ws.max_row + 1):
            if (str(ws.cell(row, idx_eval).value) == str(eval_id) and
                str(ws.cell(row, idx_activo).value) == str(activo_id)):
                ws.cell(row, idx_fecha).value = fecha
        
        guardar_workbook_atomico(wb)
