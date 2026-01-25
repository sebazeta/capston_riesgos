"""
Servicio de base de datos SQLite - Reemplazo robusto de Excel
SQLite maneja concurrencia, transacciones ACID y no se corrompe
"""
import sqlite3
import pandas as pd
import datetime as dt
import os
import threading
from typing import List, Dict, Any, Optional
from contextlib import contextmanager

DB_PATH = "tita_database.db"

# Lock global para operaciones de escritura
_db_lock = threading.Lock()


@contextmanager
def get_connection():
    """Context manager para conexiones a la base de datos"""
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.close()


def init_database():
    """Inicializa la base de datos con todas las tablas"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # EVALUACIONES
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS EVALUACIONES (
                ID_Evaluacion TEXT PRIMARY KEY,
                Nombre TEXT NOT NULL,
                Fecha TEXT,
                Estado TEXT DEFAULT 'Activa',
                Descripcion TEXT
            )
        ''')
        
        # INVENTARIO_ACTIVOS
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS INVENTARIO_ACTIVOS (
                ID_Activo TEXT PRIMARY KEY,
                ID_Evaluacion TEXT,
                Nombre_Activo TEXT NOT NULL,
                Tipo_Activo TEXT,
                Ubicacion TEXT,
                Propietario TEXT,
                Tipo_Servicio TEXT,
                App_Critica TEXT,
                Estado TEXT DEFAULT 'Pendiente',
                Fecha_Creacion TEXT,
                FOREIGN KEY (ID_Evaluacion) REFERENCES EVALUACIONES(ID_Evaluacion)
            )
        ''')
        
        # CRITERIOS_MAGERIT
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS CRITERIOS_MAGERIT (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                Dimension TEXT,
                "Nivel(1-5)" INTEGER,
                Descripcion TEXT,
                Ejemplo TEXT
            )
        ''')
        
        # CATALOGO_AMENAZAS_MAGERIT
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS CATALOGO_AMENAZAS_MAGERIT (
                Cod_MAGERIT TEXT PRIMARY KEY,
                Categoria TEXT,
                Amenaza TEXT,
                Descripcion TEXT,
                "Dimension(D/I/C)" TEXT,
                "Severidad_Base(1-5)" INTEGER
            )
        ''')
        
        # CATALOGO_ISO27002_2022
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS CATALOGO_ISO27002_2022 (
                Control TEXT PRIMARY KEY,
                Nombre TEXT,
                Dominio TEXT,
                Descripcion TEXT
            )
        ''')
        
        # BANCO_PREGUNTAS_FISICAS
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS BANCO_PREGUNTAS_FISICAS (
                ID_Pregunta TEXT PRIMARY KEY,
                Tipo_Activo TEXT,
                Bloque TEXT,
                Dimension TEXT,
                Pregunta TEXT,
                Opcion_1 TEXT,
                Opcion_2 TEXT,
                Opcion_3 TEXT,
                Opcion_4 TEXT,
                Peso INTEGER
            )
        ''')
        
        # BANCO_PREGUNTAS_VIRTUALES
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS BANCO_PREGUNTAS_VIRTUALES (
                ID_Pregunta TEXT PRIMARY KEY,
                Tipo_Activo TEXT,
                Bloque TEXT,
                Dimension TEXT,
                Pregunta TEXT,
                Opcion_1 TEXT,
                Opcion_2 TEXT,
                Opcion_3 TEXT,
                Opcion_4 TEXT,
                Peso INTEGER
            )
        ''')
        
        # CUESTIONARIOS
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS CUESTIONARIOS (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_Evaluacion TEXT,
                ID_Activo TEXT,
                Fecha_Version TEXT,
                ID_Pregunta TEXT,
                Bloque TEXT,
                Dimension TEXT,
                Pregunta TEXT,
                Opcion_1 TEXT,
                Opcion_2 TEXT,
                Opcion_3 TEXT,
                Opcion_4 TEXT,
                Peso INTEGER,
                FOREIGN KEY (ID_Evaluacion) REFERENCES EVALUACIONES(ID_Evaluacion),
                FOREIGN KEY (ID_Activo) REFERENCES INVENTARIO_ACTIVOS(ID_Activo)
            )
        ''')
        
        # RESPUESTAS
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS RESPUESTAS (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_Evaluacion TEXT,
                ID_Activo TEXT,
                Fecha_Cuestionario TEXT,
                ID_Pregunta TEXT,
                Bloque TEXT,
                Pregunta TEXT,
                Respuesta TEXT,
                Valor_Numerico INTEGER,
                Peso INTEGER,
                Dimension TEXT,
                Fecha TEXT,
                FOREIGN KEY (ID_Evaluacion) REFERENCES EVALUACIONES(ID_Evaluacion),
                FOREIGN KEY (ID_Activo) REFERENCES INVENTARIO_ACTIVOS(ID_Activo)
            )
        ''')
        
        # IMPACTO_ACTIVOS
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS IMPACTO_ACTIVOS (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_Evaluacion TEXT,
                ID_Activo TEXT,
                Fecha TEXT,
                Impacto_D INTEGER,
                Impacto_I INTEGER,
                Impacto_C INTEGER,
                Justificacion_D TEXT,
                Justificacion_I TEXT,
                Justificacion_C TEXT,
                FOREIGN KEY (ID_Evaluacion) REFERENCES EVALUACIONES(ID_Evaluacion),
                FOREIGN KEY (ID_Activo) REFERENCES INVENTARIO_ACTIVOS(ID_Activo),
                UNIQUE(ID_Evaluacion, ID_Activo)
            )
        ''')
        
        # ANALISIS_RIESGO
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ANALISIS_RIESGO (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ID_Evaluacion TEXT,
                ID_Activo TEXT,
                Fecha TEXT,
                Tipo_Activo TEXT,
                Nombre_Activo TEXT,
                Probabilidad REAL,
                Impacto REAL,
                Riesgo_Inherente REAL,
                Nivel_Riesgo TEXT,
                Recomendaciones TEXT,
                Estado TEXT,
                Modelo_IA TEXT,
                FOREIGN KEY (ID_Evaluacion) REFERENCES EVALUACIONES(ID_Evaluacion),
                FOREIGN KEY (ID_Activo) REFERENCES INVENTARIO_ACTIVOS(ID_Activo)
            )
        ''')
        
        # Crear índices para mejor rendimiento
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_activos_eval ON INVENTARIO_ACTIVOS(ID_Evaluacion)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_cuestionarios_eval_activo ON CUESTIONARIOS(ID_Evaluacion, ID_Activo)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_respuestas_eval_activo ON RESPUESTAS(ID_Evaluacion, ID_Activo)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_impacto_eval_activo ON IMPACTO_ACTIVOS(ID_Evaluacion, ID_Activo)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_analisis_eval_activo ON ANALISIS_RIESGO(ID_Evaluacion, ID_Activo)')


def read_table(table_name: str) -> pd.DataFrame:
    """Lee una tabla como DataFrame"""
    with get_connection() as conn:
        try:
            return pd.read_sql_query(f'SELECT * FROM "{table_name}"', conn)
        except Exception as e:
            print(f"Error leyendo tabla {table_name}: {e}")
            return pd.DataFrame()


def insert_row(table_name: str, data: Dict[str, Any]):
    """Inserta una fila en una tabla"""
    with _db_lock:
        with get_connection() as conn:
            columns = ', '.join([f'"{k}"' for k in data.keys()])
            placeholders = ', '.join(['?' for _ in data])
            query = f'INSERT INTO "{table_name}" ({columns}) VALUES ({placeholders})'
            conn.execute(query, list(data.values()))


def insert_rows(table_name: str, rows: List[Dict[str, Any]]):
    """Inserta múltiples filas en una tabla"""
    if not rows:
        return
    
    with _db_lock:
        with get_connection() as conn:
            columns = ', '.join([f'"{k}"' for k in rows[0].keys()])
            placeholders = ', '.join(['?' for _ in rows[0]])
            query = f'INSERT INTO "{table_name}" ({columns}) VALUES ({placeholders})'
            for row in rows:
                conn.execute(query, list(row.values()))


def update_row(table_name: str, updates: Dict[str, Any], conditions: Dict[str, Any]):
    """
    Actualiza filas que cumplan las condiciones
    
    Args:
        table_name: Nombre de la tabla
        updates: Diccionario con columnas y valores a actualizar
        conditions: Diccionario con condiciones WHERE
    """
    with _db_lock:
        with get_connection() as conn:
            set_clause = ', '.join([f'"{k}" = ?' for k in updates.keys()])
            where_clause = ' AND '.join([f'"{k}" = ?' for k in conditions.keys()])
            query = f'UPDATE "{table_name}" SET {set_clause} WHERE {where_clause}'
            conn.execute(query, list(updates.values()) + list(conditions.values()))


def delete_row(table_name: str, conditions: Dict[str, Any]):
    """
    Elimina filas que cumplan las condiciones
    
    Args:
        table_name: Nombre de la tabla
        conditions: Diccionario con condiciones WHERE
    """
    with _db_lock:
        with get_connection() as conn:
            where_clause = ' AND '.join([f'"{k}" = ?' for k in conditions.keys()])
            query = f'DELETE FROM "{table_name}" WHERE {where_clause}'
            conn.execute(query, list(conditions.values()))


def delete_rows(table_name: str, conditions: Dict[str, Any]):
    """Elimina filas que cumplan las condiciones"""
    with _db_lock:
        with get_connection() as conn:
            where_clause = ' AND '.join([f'"{k}" = ?' for k in conditions.keys()])
            query = f'DELETE FROM "{table_name}" WHERE {where_clause}'
            conn.execute(query, list(conditions.values()))


def query_rows(table_name: str, conditions: Dict[str, Any] = None) -> pd.DataFrame:
    """Consulta filas con condiciones opcionales"""
    with get_connection() as conn:
        if conditions:
            where_clause = ' AND '.join([f'"{k}" = ?' for k in conditions.keys()])
            query = f'SELECT * FROM "{table_name}" WHERE {where_clause}'
            return pd.read_sql_query(query, conn, params=list(conditions.values()))
        else:
            return pd.read_sql_query(f'SELECT * FROM "{table_name}"', conn)


def row_exists(table_name: str, conditions: Dict[str, Any]) -> bool:
    """Verifica si existe una fila con las condiciones dadas"""
    with get_connection() as conn:
        where_clause = ' AND '.join([f'"{k}" = ?' for k in conditions.keys()])
        query = f'SELECT COUNT(*) FROM "{table_name}" WHERE {where_clause}'
        result = conn.execute(query, list(conditions.values())).fetchone()
        return result[0] > 0


def upsert_row(table_name: str, data: Dict[str, Any], key_columns: List[str]):
    """
    Inserta o actualiza una fila basándose en las columnas clave
    
    Args:
        table_name: Nombre de la tabla
        data: Diccionario con todos los datos
        key_columns: Lista de columnas que forman la clave única
    """
    conditions = {col: data[col] for col in key_columns if col in data}
    if row_exists(table_name, conditions):
        updates = {k: v for k, v in data.items() if k not in key_columns}
        update_row(table_name, updates, conditions)
    else:
        insert_row(table_name, data)


# ==================== FUNCIONES DE COMPATIBILIDAD CON EXCEL SERVICE ====================
# Estas funciones mantienen la misma interfaz que excel_service.py

def ensure_workbook():
    """Compatibilidad - Verifica que la BD existe"""
    if not os.path.exists(DB_PATH):
        init_database()
    return True


def ensure_sheet_exists(sheet_name: str, headers: list):
    """Compatibilidad - Las tablas ya existen en SQLite"""
    pass  # Las tablas se crean en init_database()


def read_sheet(sheet_name: str) -> pd.DataFrame:
    """Compatibilidad - Lee una tabla como DataFrame"""
    return read_table(sheet_name)


def append_rows(sheet_name: str, rows: List[Dict[str, Any]]):
    """Compatibilidad - Inserta filas en una tabla"""
    insert_rows(sheet_name, rows)


def set_eval_active(eval_id: str, nombre: str):
    """Establece una evaluación como activa"""
    with _db_lock:
        with get_connection() as conn:
            # Desactivar todas
            conn.execute('UPDATE EVALUACIONES SET Estado = "Inactiva"')
            
            # Verificar si existe
            result = conn.execute(
                'SELECT COUNT(*) FROM EVALUACIONES WHERE ID_Evaluacion = ?', 
                [eval_id]
            ).fetchone()
            
            now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            if result[0] == 0:
                # Insertar nueva
                conn.execute('''
                    INSERT INTO EVALUACIONES (ID_Evaluacion, Nombre, Fecha, Estado, Descripcion)
                    VALUES (?, ?, ?, 'Activa', 'Evaluación activa desde la GUI')
                ''', [eval_id, nombre, now])
            else:
                # Actualizar existente
                conn.execute('''
                    UPDATE EVALUACIONES 
                    SET Nombre = ?, Fecha = ?, Estado = 'Activa', Descripcion = 'Evaluación activa desde la GUI'
                    WHERE ID_Evaluacion = ?
                ''', [nombre, now, eval_id])


def update_cuestionarios_version(eval_id: str, activo_id: str, fecha: str):
    """Actualiza la versión del cuestionario"""
    with _db_lock:
        with get_connection() as conn:
            conn.execute('''
                UPDATE CUESTIONARIOS 
                SET Fecha_Version = ?
                WHERE ID_Evaluacion = ? AND ID_Activo = ?
            ''', [fecha, eval_id, activo_id])


# ==================== EXPORTACIÓN A EXCEL ====================

def exportar_a_excel(output_path: str = "reporte_tita.xlsx"):
    """Exporta toda la base de datos a Excel para reportes"""
    import openpyxl
    from openpyxl import Workbook
    
    wb = Workbook()
    wb.remove(wb.active)
    
    tables = [
        'EVALUACIONES', 'INVENTARIO_ACTIVOS', 'CRITERIOS_MAGERIT',
        'CATALOGO_AMENAZAS_MAGERIT', 'CATALOGO_ISO27002_2022',
        'BANCO_PREGUNTAS_FISICAS', 'BANCO_PREGUNTAS_VIRTUALES',
        'CUESTIONARIOS', 'RESPUESTAS', 'IMPACTO_ACTIVOS', 'ANALISIS_RIESGO'
    ]
    
    for table in tables:
        df = read_table(table)
        ws = wb.create_sheet(table)
        
        # Headers
        for col, header in enumerate(df.columns, 1):
            ws.cell(1, col, header)
        
        # Data
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row_idx, col_idx, value)
    
    wb.save(output_path)
    return output_path
