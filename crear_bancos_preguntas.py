"""
Script para crear bancos de preguntas - Versión Oficial
Basado en BIA, MAGERIT, Continuidad del Negocio y Ciberseguridad
21 preguntas por banco con opciones de 4 niveles
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_PATH = "matriz_riesgos_v2.xlsx"

print("📝 Creando bancos de preguntas oficiales...")

# ============ BANCO 1: SERVIDOR FÍSICO (21 preguntas) ============
preguntas_fisicas = [
    ["ID_Pregunta", "Tipo_Activo", "Bloque", "Dimension", "Pregunta", "Opcion_1", "Opcion_2", "Opcion_3", "Opcion_4", "Peso"],
    
    # BLOQUE A — IMPACTO (BIA / DIC)
    ["PF-A-001", "Servidor Físico", "A-Impacto", "D", "¿Qué tan crítico es este servidor para la operación institucional?", "Bajo", "Medio", "Alto", "Crítico", 5],
    ["PF-A-002", "Servidor Físico", "A-Impacto", "D", "¿Cuántos procesos o servicios institucionales dependen directamente de este servidor?", "1", "2–3", "4–6", "7 o más", 4],
    ["PF-A-003", "Servidor Físico", "A-Impacto", "D", "Impacto si el servidor deja de estar disponible.", "Bajo", "Medio", "Alto", "Crítico", 5],
    ["PF-A-004", "Servidor Físico", "A-Impacto", "I", "Impacto si se altera la integridad de la información alojada.", "Bajo", "Medio", "Alto", "Crítico", 5],
    ["PF-A-005", "Servidor Físico", "A-Impacto", "C", "Impacto si se expone información confidencial.", "Bajo", "Medio", "Alto", "Crítico", 5],
    
    # BLOQUE B — RTO / RPO (CONTINUIDAD)
    ["PF-B-001", "Servidor Físico", "B-Continuidad", "D", "¿Cuál es el RTO esperado para este servidor (tiempo máximo tolerable para restablecer el servicio)?", "< 1 hora", "1–4 horas", "4–24 horas", "> 24 horas", 5],
    ["PF-B-002", "Servidor Físico", "B-Continuidad", "I", "¿Cuál es el RPO esperado (pérdida máxima tolerable de datos)?", "0 (sin pérdida)", "< 1 hora", "1–24 horas", "> 24 horas", 5],
    ["PF-B-003", "Servidor Físico", "B-Continuidad", "D", "¿Existen procedimientos documentados de recuperación del servidor?", "No", "Parcial", "Sí", "Sí y probados", 4],
    ["PF-B-004", "Servidor Físico", "B-Continuidad", "D", "¿El personal interno puede ejecutar la recuperación sin depender de terceros?", "No", "Depende de terceros", "Sí con apoyo", "Sí completamente", 4],
    
    # BLOQUE C — CONTROLES OPERATIVOS
    ["PF-C-001", "Servidor Físico", "C-Controles", "D", "¿Existe redundancia física (discos, fuentes, enlaces, etc.)?", "No", "Básica", "Media", "Alta", 5],
    ["PF-C-002", "Servidor Físico", "C-Controles", "D", "¿Existe protección eléctrica adecuada (UPS, planta, pruebas)?", "No", "UPS básica", "UPS + respaldo", "UPS + planta + pruebas", 5],
    ["PF-C-003", "Servidor Físico", "C-Controles", "D", "¿El ambiente físico está controlado (temperatura, humedad, acceso)?", "No", "Parcial", "Sí", "Sí + monitoreo", 4],
    ["PF-C-004", "Servidor Físico", "C-Controles", "I", "¿Se realizan respaldos periódicos de la información?", "No", "Ocasional", "Periódico", "Automático y monitoreado", 5],
    ["PF-C-005", "Servidor Físico", "C-Controles", "I", "¿Se prueban regularmente las restauraciones de respaldos?", "Nunca", "Esporádico", "Periódico", "Periódico y documentado", 5],
    
    # BLOQUE D — CIBERSEGURIDAD
    ["PF-D-001", "Servidor Físico", "D-Ciberseguridad", "C", "¿El acceso físico al servidor está controlado y auditado?", "No", "Parcial", "Sí", "Sí + registro/auditoría", 5],
    ["PF-D-002", "Servidor Físico", "D-Ciberseguridad", "C", "¿El acceso administrativo requiere autenticación fuerte (roles definidos, MFA, segregación)?", "No", "Parcial", "Sí", "Sí + segregación formal", 5],
    ["PF-D-003", "Servidor Físico", "D-Ciberseguridad", "I", "¿Se aplican parches y actualizaciones con control de cambios?", "No", "Irregular", "Regular", "Regular y auditado", 4],
    ["PF-D-004", "Servidor Físico", "D-Ciberseguridad", "C", "¿Existen registros (logs) de seguridad y monitoreo de eventos?", "No", "Locales", "Centralizados", "Centralizados + alertas", 4],
    
    # BLOQUE E — HISTORIAL / EXPOSICIÓN
    ["PF-E-001", "Servidor Físico", "E-Exposición", "C", "¿El servidor está expuesto a redes externas (Internet o DMZ)?", "No", "Exposición limitada", "Expuesto", "Expuesto y crítico", 5],
    ["PF-E-002", "Servidor Físico", "E-Exposición", "D", "¿Ha presentado incidentes de seguridad o fallas graves en el último año?", "No", "1 vez", "2–3 veces", "Frecuente", 4],
    ["PF-E-003", "Servidor Físico", "E-Exposición", "D", "¿El riesgo residual actual es aceptable para la institución?", "Sí", "Sí con observaciones", "No", "No (acción urgente)", 5],
]

# ============ BANCO 2: SERVIDOR VIRTUAL (21 preguntas) ============
preguntas_virtuales = [
    ["ID_Pregunta", "Tipo_Activo", "Bloque", "Dimension", "Pregunta", "Opcion_1", "Opcion_2", "Opcion_3", "Opcion_4", "Peso"],
    
    # BLOQUE A — IMPACTO (BIA / DIC)
    ["PV-A-001", "Servidor Virtual", "A-Impacto", "D", "¿Qué tan crítico es este servidor virtual para la operación institucional?", "Bajo", "Medio", "Alto", "Crítico", 5],
    ["PV-A-002", "Servidor Virtual", "A-Impacto", "D", "¿Cuántos servicios o aplicaciones dependen de esta máquina virtual?", "1", "2–3", "4–6", "7 o más", 4],
    ["PV-A-003", "Servidor Virtual", "A-Impacto", "D", "Impacto si la máquina virtual queda indisponible.", "Bajo", "Medio", "Alto", "Crítico", 5],
    ["PV-A-004", "Servidor Virtual", "A-Impacto", "I", "Impacto si se altera la integridad de la información o configuración de la VM.", "Bajo", "Medio", "Alto", "Crítico", 5],
    ["PV-A-005", "Servidor Virtual", "A-Impacto", "C", "Impacto si se expone información confidencial desde la VM.", "Bajo", "Medio", "Alto", "Crítico", 5],
    
    # BLOQUE B — RTO / RPO (CONTINUIDAD)
    ["PV-B-001", "Servidor Virtual", "B-Continuidad", "D", "¿Cuál es el RTO esperado para esta VM (tiempo máximo tolerable para restablecer el servicio)?", "< 1 hora", "1–4 horas", "4–24 horas", "> 24 horas", 5],
    ["PV-B-002", "Servidor Virtual", "B-Continuidad", "I", "¿Cuál es el RPO esperado para los datos o servicios de la VM (pérdida máxima tolerable)?", "0 (sin pérdida)", "< 1 hora", "1–24 horas", "> 24 horas", 5],
    ["PV-B-003", "Servidor Virtual", "B-Continuidad", "D", "¿Existe procedimiento documentado de recuperación (restauración, rebuild, snapshots)?", "No", "Parcial", "Sí", "Sí y probado", 4],
    ["PV-B-004", "Servidor Virtual", "B-Continuidad", "D", "¿La recuperación puede ejecutarse sin depender del proveedor externo?", "No", "Depende", "Sí con apoyo", "Sí completamente", 4],
    
    # BLOQUE C — CONTROLES DE PLATAFORMA
    ["PV-C-001", "Servidor Virtual", "C-Controles", "D", "¿La plataforma de virtualización cuenta con alta disponibilidad o failover?", "No", "Básica", "Media", "Alta", 5],
    ["PV-C-002", "Servidor Virtual", "C-Controles", "D", "¿El almacenamiento donde reside la VM tiene redundancia o replicación?", "No", "Básica", "Redundante", "Redundante + replicado", 5],
    ["PV-C-003", "Servidor Virtual", "C-Controles", "D", "¿La VM dispone de recursos garantizados (CPU, RAM, almacenamiento)?", "No", "Parcial", "Garantizados", "Garantizados + monitoreo", 4],
    ["PV-C-004", "Servidor Virtual", "C-Controles", "I", "¿Existen respaldos periódicos de la VM o de sus datos?", "No", "Ocasional", "Periódico", "Automático y monitoreado", 5],
    ["PV-C-005", "Servidor Virtual", "C-Controles", "I", "¿Se prueban regularmente las restauraciones de la VM o de los datos?", "Nunca", "Esporádico", "Periódico", "Periódico y documentado", 5],
    
    # BLOQUE D — CIBERSEGURIDAD
    ["PV-D-001", "Servidor Virtual", "D-Ciberseguridad", "C", "¿La VM tiene hardening aplicado (servicios mínimos, firewall, baseline)?", "No", "Parcial", "Sí", "Sí + baseline formal", 5],
    ["PV-D-002", "Servidor Virtual", "D-Ciberseguridad", "C", "¿El acceso administrativo a la VM está protegido (roles, MFA, segregación)?", "No", "Parcial", "Sí", "Sí + segregación formal", 5],
    ["PV-D-003", "Servidor Virtual", "D-Ciberseguridad", "I", "¿Se gestionan parches y actualizaciones con control de cambios?", "No", "Irregular", "Regular", "Regular y auditado", 4],
    ["PV-D-004", "Servidor Virtual", "D-Ciberseguridad", "C", "¿Existen logs de seguridad y monitoreo centralizado de eventos?", "No", "Locales", "Centralizados", "Centralizados + alertas", 4],
    
    # BLOQUE E — EXPOSICIÓN / HISTORIAL
    ["PV-E-001", "Servidor Virtual", "E-Exposición", "C", "¿La VM está expuesta a Internet o zonas desmilitarizadas (DMZ)?", "No", "Exposición limitada", "Expuesta", "Expuesta y crítica", 5],
    ["PV-E-002", "Servidor Virtual", "E-Exposición", "D", "¿La VM ha tenido incidentes de seguridad o caídas en el último año?", "No", "1 vez", "2–3 veces", "Frecuente", 4],
    ["PV-E-003", "Servidor Virtual", "E-Exposición", "D", "¿El riesgo residual actual es aceptable para la institución?", "Sí", "Sí con observaciones", "No", "No (acción urgente)", 5],
]

# ============ CREAR/ACTUALIZAR EXCEL ============
try:
    wb = load_workbook(EXCEL_PATH)
    
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    bloque_fills = {
        "A-Impacto": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        "B-Continuidad": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        "C-Controles": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "D-Ciberseguridad": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
        "E-Exposición": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
    }
    
    # ===== BANCO PREGUNTAS FÍSICAS =====
    if "BANCO_PREGUNTAS_FISICAS" in wb.sheetnames:
        del wb["BANCO_PREGUNTAS_FISICAS"]
    
    ws_fis = wb.create_sheet("BANCO_PREGUNTAS_FISICAS")
    for row in preguntas_fisicas:
        ws_fis.append(row)
    
    for cell in ws_fis[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for row_idx in range(2, len(preguntas_fisicas) + 1):
        bloque = ws_fis.cell(row=row_idx, column=3).value
        if bloque in bloque_fills:
            for col in range(1, 11):
                ws_fis.cell(row=row_idx, column=col).fill = bloque_fills[bloque]
    
    ws_fis.column_dimensions['A'].width = 12
    ws_fis.column_dimensions['B'].width = 15
    ws_fis.column_dimensions['C'].width = 15
    ws_fis.column_dimensions['D'].width = 5
    ws_fis.column_dimensions['E'].width = 70
    ws_fis.column_dimensions['F'].width = 20
    ws_fis.column_dimensions['G'].width = 20
    ws_fis.column_dimensions['H'].width = 20
    ws_fis.column_dimensions['I'].width = 25
    ws_fis.column_dimensions['J'].width = 6
    
    print("✅ BANCO_PREGUNTAS_FISICAS creado (21 preguntas)")
    
    # ===== BANCO PREGUNTAS VIRTUALES =====
    if "BANCO_PREGUNTAS_VIRTUALES" in wb.sheetnames:
        del wb["BANCO_PREGUNTAS_VIRTUALES"]
    
    ws_vir = wb.create_sheet("BANCO_PREGUNTAS_VIRTUALES")
    for row in preguntas_virtuales:
        ws_vir.append(row)
    
    for cell in ws_vir[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for row_idx in range(2, len(preguntas_virtuales) + 1):
        bloque = ws_vir.cell(row=row_idx, column=3).value
        if bloque in bloque_fills:
            for col in range(1, 11):
                ws_vir.cell(row=row_idx, column=col).fill = bloque_fills[bloque]
    
    ws_vir.column_dimensions['A'].width = 12
    ws_vir.column_dimensions['B'].width = 15
    ws_vir.column_dimensions['C'].width = 15
    ws_vir.column_dimensions['D'].width = 5
    ws_vir.column_dimensions['E'].width = 70
    ws_vir.column_dimensions['F'].width = 20
    ws_vir.column_dimensions['G'].width = 20
    ws_vir.column_dimensions['H'].width = 20
    ws_vir.column_dimensions['I'].width = 25
    ws_vir.column_dimensions['J'].width = 6
    
    print("✅ BANCO_PREGUNTAS_VIRTUALES creado (21 preguntas)")
    
    wb.save(EXCEL_PATH)
    print(f"\n🎉 Bancos de preguntas oficiales creados")
    print(f"   - {len(preguntas_fisicas)-1} preguntas Servidores Físicos")
    print(f"   - {len(preguntas_virtuales)-1} preguntas Servidores Virtuales")
    print(f"\n📋 Bloques: A-Impacto, B-Continuidad, C-Controles, D-Ciberseguridad, E-Exposición")

except Exception as e:
    print(f"❌ Error: {e}")
