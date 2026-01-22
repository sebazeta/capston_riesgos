import json, re
import datetime as dt
import pandas as pd
import streamlit as st
import requests
from openpyxl import load_workbook, Workbook

EXCEL_PATH = "matriz_riesgos_v2.xlsx"
OLLAMA_URL = "http://localhost:11434/api/generate"

st.set_page_config(page_title="Matriz MAGERIT + Ollama (v3)", layout="wide")
st.title("Capstone: Matriz de Riesgos (MAGERIT + ISO27002:2022) + Ollama")
st.sidebar.write("📄 Excel en uso:", EXCEL_PATH)


# -------------------- HEADERS --------------------
CUESTIONARIOS_HEADERS = [
    "ID_Evaluacion",
    "ID_Activo",
    "Fecha",                # versión del cuestionario (timestamp)
    "ID_Pregunta",
    "Pregunta",
    "Tipo_Respuesta",
    "Peso",
    "Dimension",            # D/I/C
    "Fuente(IA/Base)"
]

RESPUESTAS_HEADERS = [
    "ID_Evaluacion",
    "ID_Activo",
    "Fecha_Cuestionario",   # versión respondida (timestamp)
    "ID_Pregunta",
    "Pregunta",
    "Respuesta",
    "Tipo_Respuesta",
    "Peso",
    "Dimension",
    "Fecha"                 # cuándo respondió
]

IMPACTO_HEADERS = [
    "ID_Evaluacion",
    "ID_Activo",
    "Fecha_Cuestionario",
    "Impacto_D",
    "Impacto_I",
    "Impacto_C",
    "Impacto_Global",
    "Fecha_Calculo"
]


# -------------------- EXCEL HELPERS --------------------
def ensure_workbook():
    try:
        wb = load_workbook(EXCEL_PATH)
        return wb
    except FileNotFoundError:
        wb = Workbook()
        ws0 = wb.active
        ws0.title = "INVENTARIO_ACTIVOS"
        ws0.append(["ID_Activo", "Nombre", "Tipo_Activo", "Propietario", "Ubicacion"])
        wb.save(EXCEL_PATH)
        return load_workbook(EXCEL_PATH)


def ensure_sheet_exists(sheet_name: str, headers: list):
    wb = ensure_workbook()
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        wb.save(EXCEL_PATH)


def read_sheet(sheet_name: str) -> pd.DataFrame:
    wb = ensure_workbook()
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    return pd.read_excel(EXCEL_PATH, sheet_name=sheet_name)


def append_rows(sheet_name: str, rows: list):
    if not rows:
        return
    wb = ensure_workbook()
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name} no existe. Crea con ensure_sheet_exists() primero.")
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    for r in rows:
        ws.append([r.get(h) for h in headers])
    wb.save(EXCEL_PATH)


def set_eval_active(eval_id: str, nombre: str):
    ensure_sheet_exists("EVALUACIONES", ["ID_Evaluacion", "Nombre", "Fecha", "Estado", "Descripcion"])
    wb = load_workbook(EXCEL_PATH)
    ws = wb["EVALUACIONES"]
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    # set all inactive
    for r in range(2, ws.max_row + 1):
        ws.cell(r, idx.get("Estado", 4)).value = "Inactiva"

    found = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, idx["ID_Evaluacion"]).value) == str(eval_id):
            found = r
            break

    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = {
        "ID_Evaluacion": eval_id,
        "Nombre": nombre,
        "Fecha": now,
        "Estado": "Activa",
        "Descripcion": "Evaluación activa desde la GUI"
    }

    if found is None:
        ws.append([row.get(h) for h in headers])
    else:
        for h in headers:
            ws.cell(found, idx[h]).value = row.get(h)

    wb.save(EXCEL_PATH)


def update_cuestionarios_version(eval_id: str, activo_id: str, fecha_version: str, df_edit: pd.DataFrame):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["CUESTIONARIOS"]
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    df_edit = df_edit.copy()
    df_edit["ID_Pregunta"] = df_edit["ID_Pregunta"].astype(str)
    by_id = {str(r["ID_Pregunta"]): r for _, r in df_edit.iterrows()}

    editable = ["Pregunta", "Tipo_Respuesta", "Peso", "Dimension"]
    editable = [c for c in editable if c in df_edit.columns and c in idx]

    updated = 0
    for r in range(2, ws.max_row + 1):
        ev = str(ws.cell(r, idx["ID_Evaluacion"]).value)
        ac = str(ws.cell(r, idx["ID_Activo"]).value)
        fv = str(ws.cell(r, idx["Fecha"]).value)
        pid = str(ws.cell(r, idx["ID_Pregunta"]).value)

        if ev == str(eval_id) and ac == str(activo_id) and fv == str(fecha_version) and pid in by_id:
            row_data = by_id[pid]
            for col in editable:
                val = row_data.get(col)

                if col == "Tipo_Respuesta":
                    v = str(val).strip().lower()
                    if "0/1" in v or "si/no" in v or "sí/no" in v:
                        val = "0/1"
                    elif "1-5" in v:
                        val = "1-5"

                if col == "Peso":
                    try:
                        val = int(val)
                    except:
                        val = 3
                    val = max(1, min(5, val))

                if col == "Dimension":
                    v = str(val).strip().upper()
                    if v not in ["D", "I", "C"]:
                        v = "I"
                    val = v

                ws.cell(r, idx[col]).value = val

            updated += 1

    wb.save(EXCEL_PATH)
    return updated


# -------------------- OLLAMA + PARSING --------------------
def ollama_generate(model: str, prompt: str):
    try:
        r = requests.post(
            OLLAMA_URL,
            json={
                "model": model,
                "prompt": prompt,
                "stream": False,
                "options": {
                    "num_predict": 350,
                    "temperature": 0.2
                }
            },
            timeout=90
        )
        r.raise_for_status()
        return r.json().get("response", "")
    except requests.exceptions.Timeout:
        return "__TIMEOUT__"
    except Exception as e:
        return f"__ERROR__:{str(e)}"


def extract_json_array(text: str):
    try:
        match = re.search(r"\[.*\]", text, re.DOTALL)
        if not match:
            return []
        arr = match.group(0)
        arr = re.sub(r",\s*}", "}", arr)
        arr = re.sub(r",\s*\]", "]", arr)
        return json.loads(arr)
    except Exception:
        return []


def validate_ia_questions(qs, n_ia: int):
    cleaned = []
    if not isinstance(qs, list):
        return cleaned

    for item in qs:
        if not isinstance(item, dict):
            continue

        pregunta = item.get("Pregunta") or item.get("pregunta") or item.get("question")
        tr = item.get("Tipo_Respuesta") or item.get("tipo_respuesta") or item.get("answer_type")
        peso = item.get("Peso") or item.get("weight", 3)
        dim = item.get("Dimension") or item.get("dimension", "I")

        if not pregunta or tr is None:
            continue

        t = str(tr).strip().lower()
        if "0/1" in t or "0-1" in t or "si/no" in t or "sí/no" in t:
            tr = "0/1"
        elif "1-5" in t:
            tr = "1-5"
        else:
            if str(tr).strip().upper() in ["D", "I", "C"]:
                tr = "1-5"
            else:
                continue

        try:
            peso = int(peso)
        except:
            peso = 3
        peso = max(1, min(5, peso))

        dim = str(dim).strip().upper()
        if dim not in ["D", "I", "C"]:
            dim = "I"

        cleaned.append({
            "Pregunta": str(pregunta).strip(),
            "Tipo_Respuesta": tr,
            "Peso": peso,
            "Dimension": dim
        })

        if len(cleaned) >= n_ia:
            break

    return cleaned


FALLBACK_JSON = """
[
  {"Pregunta":"¿El sistema cuenta con redundancia que permita cumplir el RTO objetivo del activo?","Tipo_Respuesta":"1-5","Peso":4,"Dimension":"D"},
  {"Pregunta":"¿Los respaldos y replicación permiten minimizar la pérdida de datos dentro del umbral esperado?","Tipo_Respuesta":"1-5","Peso":4,"Dimension":"I"},
  {"Pregunta":"¿Hay controles para mantener la confidencialidad de datos durante contingencia y restauración?","Tipo_Respuesta":"1-5","Peso":3,"Dimension":"C"}
]
"""


# -------------------- ENSURE SHEETS --------------------
ensure_sheet_exists("CUESTIONARIOS", CUESTIONARIOS_HEADERS)
ensure_sheet_exists("RESPUESTAS", RESPUESTAS_HEADERS)
ensure_sheet_exists("IMPACTO_ACTIVOS", IMPACTO_HEADERS)


# -------------------- UI TABS --------------------
tab1, tab2, tab3, tab4 = st.tabs([
    "Inventario",
    "Preguntas (IA+BIA)",
    "Responder cuestionario",
    "Cálculo de Impacto (MAGERIT)"
])

# -------- TAB 1 --------
with tab1:
    st.subheader("Inventario de activos")
    inv = read_sheet("INVENTARIO_ACTIVOS")
    if inv.empty:
        st.warning("No hay activos. Crea al menos uno en la hoja INVENTARIO_ACTIVOS.")
    else:
        st.dataframe(inv, use_container_width=True)


# -------- TAB 2 --------
with tab2:
    st.subheader("Generar cuestionario (versionado por Fecha)")

    inv = read_sheet("INVENTARIO_ACTIVOS")
    bank = read_sheet("BANCO_PREGUNTAS")
    cu = read_sheet("CUESTIONARIOS")

    if inv.empty:
        st.warning("No hay activos en INVENTARIO_ACTIVOS.")
        st.stop()
    if bank.empty:
        st.warning("BANCO_PREGUNTAS está vacío (corre tu seed_catalogos.py).")
        st.stop()

    c1, c2, c3 = st.columns(3)
    with c1:
        activo_id = st.selectbox("Activo", inv["ID_Activo"].tolist(), key="t2_activo")
    with c2:
        eval_id = st.text_input("ID Evaluación", value="EVA-001", key="t2_eval_id")
    with c3:
        model = st.selectbox("Modelo Ollama", ["llama3", "phi3", "mistral"], index=0, key="t2_model")

    eval_name = st.text_input("Nombre Evaluación", value="Evaluación desde GUI", key="t2_eval_name")

    cu_act = cu[
        (cu["ID_Evaluacion"].astype(str) == str(eval_id)) &
        (cu["ID_Activo"].astype(str) == str(activo_id))
    ].copy() if not cu.empty else pd.DataFrame()

    versiones = sorted(cu_act["Fecha"].dropna().astype(str).unique().tolist()) if (not cu_act.empty and "Fecha" in cu_act.columns) else []
    ultima_version = versiones[-1] if versiones else None

    st.markdown("### 📌 Estado del cuestionario para este activo")
    box1, box2, box3 = st.columns(3)
    box1.metric("Total versiones", len(versiones))
    box2.metric("Última versión", ultima_version if ultima_version else "—")
    box3.metric("Preguntas última versión", int(len(cu_act[cu_act["Fecha"].astype(str) == str(ultima_version)])) if ultima_version else 0)

    st.markdown("---")

    if not versiones:
        opciones = ["Generar primera versión"]
    else:
        opciones = ["Usar la última versión (no generar)", "Generar nueva versión (crear otra Fecha)"]

    accion = st.radio("¿Qué quieres hacer?", opciones, index=0, key="t2_accion")

    n_base = 5
    n_ia = 15

    if st.button("🚀 Ejecutar", key="t2_btn_run"):

        # ====== CASO A: USAR ÚLTIMA VERSIÓN ======
        if "Usar la última versión" in accion:
            st.session_state["go_to_tab3"] = True
            st.session_state["resp_eval"] = str(eval_id)
            st.session_state["resp_act"] = str(activo_id)
            st.session_state["resp_fecha"] = str(ultima_version)

            st.success(f"✅ No se generó nada. Usarás la versión existente en Tab3.\n\nÚltima versión: **{ultima_version}**")
            st.info("👉 Ve a **Tab 3 (Responder cuestionario)**: se precargará Evaluación/Activo/Versión.")
            # IMPORTANTÍSIMO: NO usamos st.stop() (porque mata Tab3/Tab4 y quedan en negro)
            st.rerun()

        # ====== CASO B: GENERAR PRIMERA o NUEVA VERSIÓN ======
        fecha_version = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        arow = inv[inv["ID_Activo"].astype(str) == str(activo_id)].iloc[0].to_dict()
        tipo = str(arow.get("Tipo_Activo", "General")).lower()

        # --------------- BASE ---------------
        b = bank.copy()
        b["Tipo_Activo"] = b.get("Tipo_Activo", "").fillna("").astype(str)
        match = b[b["Tipo_Activo"].str.lower() == tipo]
        base_sel = match.head(n_base)

        if len(base_sel) < n_base:
            base_sel = pd.concat(
                [base_sel, b[~b.index.isin(base_sel.index)].head(n_base - len(base_sel))],
                ignore_index=True
            )

        base_rows = []
        for _, r in base_sel.iterrows():
            base_rows.append({
                "ID_Evaluacion": str(eval_id),
                "ID_Activo": str(activo_id),
                "Fecha": fecha_version,
                "ID_Pregunta": str(r.get("ID_Pregunta", f"P-{dt.datetime.now().strftime('%H%M%S')}")),
                "Pregunta": str(r.get("Pregunta", "")),
                "Tipo_Respuesta": str(r.get("Tipo_Respuesta(0/1|1-5)", "0/1")).replace(" ", ""),
                "Peso": int(r.get("Peso(1-5)", 3)),
                "Dimension": str(r.get("Dimension", "I")).strip().upper() if "Dimension" in r else "I",
                "Fuente(IA/Base)": "Base"
            })

        # --------------- IA ---------------
        prompt = f"""
Devuelve SOLO un JSON válido. NO texto adicional. NO explicación.
Genera EXACTAMENTE {n_ia} preguntas TÉCNICAS para continuidad (RTO/RPO/BIA aplicados al activo):
- arquitectura, HA, redundancia, backups, replicación, dependencias, restauración, pruebas.
- NO preguntes literalmente "¿Cuál es el RTO/RPO?".
Formato obligatorio:
[
  {{
    "Pregunta": "texto",
    "Tipo_Respuesta": "0/1" o "1-5",
    "Peso": 1-5,
    "Dimension": "D" o "I" o "C"
  }}
]
Contexto del activo:
{json.dumps(arow, ensure_ascii=False)}
""".strip()

        with st.spinner("🧠 Ollama está generando preguntas..."):
            raw = ollama_generate(model, prompt)

        if raw == "__TIMEOUT__" or (isinstance(raw, str) and raw.startswith("__ERROR__")):
            st.error("⚠️ Ollama falló (timeout/error). Usando fallback para que el pipeline no se caiga.")
            raw = FALLBACK_JSON

        st.text_area("🧪 Respuesta cruda (debug)", raw, height=250)

        parsed = extract_json_array(raw)
        ia_valid = validate_ia_questions(parsed, n_ia=n_ia)
        st.success(f"✅ Preguntas IA válidas: {len(ia_valid)}")

        ia_rows = []
        for i, q in enumerate(ia_valid, start=1):
            ia_rows.append({
                "ID_Evaluacion": str(eval_id),
                "ID_Activo": str(activo_id),
                "Fecha": fecha_version,
                "ID_Pregunta": f"IA-{dt.datetime.now().strftime('%H%M%S')}-{i}",
                "Pregunta": q["Pregunta"],
                "Tipo_Respuesta": q["Tipo_Respuesta"],
                "Peso": q["Peso"],
                "Dimension": q["Dimension"],
                "Fuente(IA/Base)": "IA"
            })

        append_rows("CUESTIONARIOS", base_rows + ia_rows)
        set_eval_active(str(eval_id), str(eval_name))

        st.success(f"📥 Cuestionario guardado. Nueva versión creada: {fecha_version}")
        st.rerun()


# -------- TAB 3 --------
with tab3:
    st.subheader("Responder cuestionario (por versión/Fecha)")

    cu = read_sheet("CUESTIONARIOS")
    if cu.empty:
        st.warning("No hay cuestionarios generados todavía. Ve a Tab2 y genera una versión.")
        st.stop()

    # ✅ Precarga real desde Tab2 (session_state)
    auto_mode = st.session_state.get("go_to_tab3", False)
    auto_eval = st.session_state.get("resp_eval", None)
    auto_act = st.session_state.get("resp_act", None)
    auto_fecha = st.session_state.get("resp_fecha", None)

    # Selectores
    evals = sorted(cu["ID_Evaluacion"].dropna().astype(str).unique().tolist())
    if not evals:
        st.warning("No hay evaluaciones.")
        st.stop()

    eval_index = evals.index(str(auto_eval)) if (auto_mode and auto_eval in evals) else 0
    eval_sel = st.selectbox("Selecciona Evaluación", evals, index=eval_index, key="resp_eval_sel")

    acts = sorted(cu[cu["ID_Evaluacion"].astype(str) == str(eval_sel)]["ID_Activo"].dropna().astype(str).unique().tolist())
    if not acts:
        st.warning("No hay activos para esta evaluación.")
        st.stop()

    act_index = acts.index(str(auto_act)) if (auto_mode and auto_act in acts) else 0
    act_sel = st.selectbox("Selecciona Activo", acts, index=act_index, key="resp_act_sel")

    fechas = sorted(
        cu[
            (cu["ID_Evaluacion"].astype(str) == str(eval_sel)) &
            (cu["ID_Activo"].astype(str) == str(act_sel))
        ]["Fecha"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    if not fechas:
        st.warning("No hay versiones de cuestionario.")
        st.stop()

    fecha_index = fechas.index(str(auto_fecha)) if (auto_mode and auto_fecha in fechas) else len(fechas) - 1
    fecha_sel = st.selectbox("Selecciona versión del cuestionario", fechas, index=fecha_index, key="resp_fecha_sel")

    # guardar versión activa
    st.session_state["fecha_sel"] = str(fecha_sel)

    # ya consumimos el auto_mode
    st.session_state["go_to_tab3"] = False

    # Cargar preguntas de esa versión
    preguntas = cu[
        (cu["ID_Evaluacion"].astype(str) == str(eval_sel)) &
        (cu["ID_Activo"].astype(str) == str(act_sel)) &
        (cu["Fecha"].astype(str) == str(fecha_sel))
    ].copy()

    if preguntas.empty:
        st.warning("No hay preguntas para esa versión.")
        st.stop()

    # Evitar duplicados visuales
    for col in ["Pregunta", "Tipo_Respuesta", "Fuente(IA/Base)"]:
        if col in preguntas.columns:
            preguntas[col] = preguntas[col].astype(str)

    subset_cols = [c for c in ["Pregunta", "Tipo_Respuesta", "Peso", "Dimension", "Fuente(IA/Base)"] if c in preguntas.columns]
    if subset_cols:
        preguntas = preguntas.drop_duplicates(subset=subset_cols, keep="first").reset_index(drop=True)

    st.info(f"Total de preguntas en esta versión: {len(preguntas)}")

    # ----- Editor de preguntas (opcional) -----
    st.markdown("### ✏️ Editar preguntas (opcional)")
    editar = st.toggle("Habilitar edición de preguntas de ESTA versión", value=False, key="t3_edit_toggle")
    if editar:
        cols_show = [c for c in ["ID_Pregunta", "Pregunta", "Tipo_Respuesta", "Peso", "Dimension", "Fuente(IA/Base)"] if c in preguntas.columns]
        df_edit = preguntas[cols_show].copy()
        edited = st.data_editor(df_edit, num_rows="fixed", use_container_width=True, key="t3_editor")

        if st.button("💾 Guardar edición", key="t3_save_edit"):
            updated = update_cuestionarios_version(eval_sel, act_sel, str(fecha_sel), edited)
            st.success(f"✅ Filas actualizadas: {updated}")
            st.rerun()

    st.markdown("---")
    st.markdown("### 📝 Responder cuestionario")

    respuestas_rows = []
    form_id = f"form_{eval_sel}_{act_sel}_{fecha_sel}".replace(" ", "_").replace(":", "").replace("/", "-")

    with st.form(form_id):
        for i, q in enumerate(preguntas.to_dict("records")):
            st.markdown(f"**{i+1}. {q.get('Pregunta','')}**")

            tipo_resp = str(q.get("Tipo_Respuesta", "1-5")).strip()
            dim_val = str(q.get("Dimension", "I")).strip().upper()
            if dim_val not in ["D", "I", "C"]:
                dim_val = "I"

            widget_key = f"{eval_sel}|{act_sel}|{fecha_sel}|{q.get('ID_Pregunta','NA')}|{i}"

            if tipo_resp == "0/1":
                resp_txt = st.radio("Respuesta", ["No", "Sí"], key=widget_key)
                valor = 1 if resp_txt == "Sí" else 0
            else:
                valor = st.slider("Respuesta (1–5)", 1, 5, 3, key=widget_key)

            try:
                peso_val = int(q.get("Peso", 3))
            except:
                peso_val = 3
            peso_val = max(1, min(5, peso_val))

            respuestas_rows.append({
                "ID_Evaluacion": str(eval_sel),
                "ID_Activo": str(act_sel),
                "Fecha_Cuestionario": str(fecha_sel),
                "ID_Pregunta": str(q.get("ID_Pregunta", f"Q-{i+1}")),
                "Pregunta": str(q.get("Pregunta", "")),
                "Respuesta": valor,
                "Tipo_Respuesta": tipo_resp,
                "Peso": peso_val,
                "Dimension": dim_val,
                "Fecha": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })

        submitted = st.form_submit_button("💾 Guardar respuestas")

    if submitted:
        append_rows("RESPUESTAS", respuestas_rows)
        st.success("✅ Respuestas guardadas correctamente.")


# -------- TAB 4 --------
with tab4:
    st.subheader("📊 Cálculo de Impacto por Activo (MAGERIT)")

    resp = read_sheet("RESPUESTAS")
    if resp.empty:
        st.warning("No hay respuestas registradas todavía. Responde en Tab3 y guarda.")
        st.stop()

    if "Fecha_Cuestionario" not in resp.columns:
        st.error("RESPUESTAS no tiene Fecha_Cuestionario. Revisa headers/Excel.")
        st.stop()

    for col in ["Respuesta", "Peso"]:
        if col in resp.columns:
            resp[col] = pd.to_numeric(resp[col], errors="coerce").fillna(0)

    if "Dimension" not in resp.columns:
        resp["Dimension"] = "I"

    col1, col2, col3 = st.columns(3)
    evals = sorted(resp["ID_Evaluacion"].dropna().astype(str).unique().tolist())
    with col1:
        eval_sel = st.selectbox("Evaluación", evals, key="t4_eval")

    acts = sorted(resp[resp["ID_Evaluacion"].astype(str) == str(eval_sel)]["ID_Activo"].dropna().astype(str).unique().tolist())
    if not acts:
        st.warning("No hay activos para esta evaluación.")
        st.stop()
    with col2:
        act_sel = st.selectbox("Activo", acts, key="t4_act")

    versiones = sorted(
        resp[(resp["ID_Evaluacion"].astype(str) == str(eval_sel)) & (resp["ID_Activo"].astype(str) == str(act_sel))]["Fecha_Cuestionario"]
        .dropna().astype(str).unique().tolist()
    )
    if not versiones:
        st.warning("No hay versiones en RESPUESTAS para este activo. Guarda respuestas desde Tab3.")
        st.stop()
    with col3:
        fecha_c = st.selectbox("Versión (Fecha_Cuestionario)", versiones, key="t4_fecha")

    data = resp[
        (resp["ID_Evaluacion"].astype(str) == str(eval_sel)) &
        (resp["ID_Activo"].astype(str) == str(act_sel)) &
        (resp["Fecha_Cuestionario"].astype(str) == str(fecha_c))
    ].copy()

    if data.empty:
        st.warning("No hay respuestas para esa versión.")
        st.stop()

    data["Valor_Ponderado"] = data["Respuesta"] * data["Peso"]

    impactos = {}
    for dim in ["D", "I", "C"]:
        d = data[data["Dimension"].astype(str).str.upper().str.strip() == dim]
        impactos[dim] = round(float(d["Valor_Ponderado"].mean()), 2) if not d.empty else 0.0

    impacto_global = max(impactos.values())

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Disponibilidad (D)", impactos["D"])
    m2.metric("Integridad (I)", impactos["I"])
    m3.metric("Confidencialidad (C)", impactos["C"])
    m4.metric("Impacto Global", impacto_global)

    if st.button("💾 Guardar impacto", key="t4_save"):
        append_rows("IMPACTO_ACTIVOS", [{
            "ID_Evaluacion": str(eval_sel),
            "ID_Activo": str(act_sel),
            "Fecha_Cuestionario": str(fecha_c),
            "Impacto_D": impactos["D"],
            "Impacto_I": impactos["I"],
            "Impacto_C": impactos["C"],
            "Impacto_Global": impacto_global,
            "Fecha_Calculo": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }])
        st.success("✅ Impacto guardado en IMPACTO_ACTIVOS.")
