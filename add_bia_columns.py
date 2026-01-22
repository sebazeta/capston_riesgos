from openpyxl import load_workbook

EXCEL_PATH = "matriz_riesgos_v2.xlsx"

def add_columns(sheet, new_cols):
    headers = [c.value for c in sheet[1]]
    for col in new_cols:
        if col not in headers:
            headers.append(col)
    # reescribir encabezados
    for i, h in enumerate(headers, start=1):
        sheet.cell(row=1, column=i).value = h

def main():
    wb = load_workbook(EXCEL_PATH)

    ws = wb["INVENTARIO_ACTIVOS"]
    add_columns(ws, ["RTO_objetivo_horas", "RPO_objetivo_horas", "BIA_impacto(1-5)"])

    wb.save(EXCEL_PATH)
    print("✅ Columnas RTO/RPO/BIA añadidas a INVENTARIO_ACTIVOS.")

if __name__ == "__main__":
    main()
