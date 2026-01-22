import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

EXCEL_PATH = "matriz_riesgos.xlsx"

SHEETS = {
    "INVENTARIO_ACTIVOS": [
        "ID_Activo", "Nombre_Activo", "Tipo_Activo", "Propietario", "Ubicacion", "Criticidad_Negocio(1-5)"
    ],
    "CUESTIONARIO_MAGERIT": [
        "ID_Activo", "Fecha",
        "Backup(0/1)", "Cifrado(0/1)", "MFA(0/1)", "Firewall(0/1)", "Monitoreo(0/1)",
        "Internet_Expuesto(0/1)", "Datos_Sensibles(0/1)"
    ],
    "VALORACION_DIC": [
        "ID_Activo", "Fecha",
        "Disponibilidad(1-5)", "Integridad(1-5)", "Confidencialidad(1-5)",
        "Just_D", "Just_I", "Just_C"
    ],
    "AMENAZAS_VULNERAB": [
        "ID_Riesgo", "ID_Activo", "Fecha",
        "Amenaza", "Vulnerabilidad", "Dimension(D/I/C)", "Severidad(1-5)"
    ],
    "ANALISIS_RIESGO": [
        "ID_Riesgo", "ID_Activo", "Fecha",
        "Probabilidad(1-5)", "Impacto(1-5)", "Riesgo_Inherente(1-25)"
    ],
    "SALVAGUARDAS": [
        "ID_Riesgo", "ID_Activo", "Fecha",
        "Control", "Descripcion", "Efectividad(0-100)"
    ],
    "RIESGO_RESIDUAL": [
        "ID_Riesgo", "ID_Activo", "Fecha",
        "Riesgo_Inherente", "Efectividad", "Riesgo_Residual"
    ],
    "DASHBOARD": [
        "Fecha", "Total_Activos", "Riesgo_Promedio", "Riesgos_Altos(>=16)", "Riesgos_Medios(9-15)", "Riesgos_Bajos(<=8)"
    ],
}

def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

def main():
    if os.path.exists(EXCEL_PATH):
        print(f"Ya existe {EXCEL_PATH}. No lo sobrescribo.")
        return

    wb = Workbook()
    wb.remove(wb.active)

    for name, headers in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        autosize(ws)

    wb.save(EXCEL_PATH)
    print(f"Creado {EXCEL_PATH} con pestañas base.")

if __name__ == "__main__":
    main()
