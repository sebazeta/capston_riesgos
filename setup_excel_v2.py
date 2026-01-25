import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

EXCEL_PATH = "matriz_riesgos_v2.xlsx"

SHEETS = {
    "PORTADA": ["Campo", "Valor"],
    "EVALUACIONES": ["ID_Evaluacion", "Nombre", "Fecha", "Estado", "Descripcion"],

    "CRITERIOS_MAGERIT": ["Dimension", "Nivel(1-5)", "Descripcion", "Ejemplo"],
    "CATALOGO_AMENAZAS_MAGERIT": ["Cod_MAGERIT", "Categoria", "Amenaza", "Descripcion", "Dimension(D/I/C)", "Severidad_Base(1-5)"],
    "CATALOGO_ISO27002_2022": ["Control", "Nombre", "Dominio", "Descripcion"],

    # Activos - columnas sincronizadas con activo_service.py
    "INVENTARIO_ACTIVOS": ["ID_Evaluacion", "ID_Activo", "Nombre_Activo", "Tipo_Activo", 
                           "Ubicacion", "Propietario", "Tipo_Servicio", "App_Critica",
                           "Estado", "Fecha_Creacion"],

    # Bancos de preguntas por tipo de activo
    "BANCO_PREGUNTAS_FISICAS": ["ID_Pregunta", "Tipo_Activo", "Bloque", "Dimension", "Pregunta", 
                                 "Opcion_1", "Opcion_2", "Opcion_3", "Opcion_4", "Peso"],
    "BANCO_PREGUNTAS_VIRTUALES": ["ID_Pregunta", "Tipo_Activo", "Bloque", "Dimension", "Pregunta", 
                                   "Opcion_1", "Opcion_2", "Opcion_3", "Opcion_4", "Peso"],

    # Cuestionarios generados para cada activo
    "CUESTIONARIOS": ["ID_Evaluacion", "ID_Activo", "Fecha_Version", "ID_Pregunta", "Bloque",
                      "Dimension", "Pregunta", "Opcion_1", "Opcion_2", "Opcion_3", "Opcion_4", "Peso"],

    # Respuestas del usuario
    "RESPUESTAS": ["ID_Evaluacion", "ID_Activo", "Fecha_Cuestionario", "ID_Pregunta", "Bloque",
                   "Pregunta", "Respuesta", "Valor_Numerico", "Peso", "Dimension", "Fecha"],

    # Impacto DIC (valoración MAGERIT)
    "IMPACTO_ACTIVOS": ["ID_Evaluacion", "ID_Activo", "Fecha", "Impacto_D", "Impacto_I", "Impacto_C",
                        "Justificacion_D", "Justificacion_I", "Justificacion_C"],

    "AMENAZAS_VULNERAB": ["ID_Evaluacion", "ID_Riesgo", "ID_Activo", "Fecha", "Cod_MAGERIT", "Amenaza",
                          "Vulnerabilidad", "Dimension(D/I/C)", "Severidad(1-5)"],

    "ANALISIS_RIESGO": ["ID_Evaluacion", "ID_Activo", "Fecha", "Tipo_Activo", "Nombre_Activo",
                        "Probabilidad", "Impacto", "Riesgo_Inherente", "Nivel_Riesgo",
                        "Recomendaciones", "Estado", "Modelo_IA"],

    "SALVAGUARDAS": ["ID_Evaluacion", "ID_Riesgo", "ID_Activo", "Fecha", "Control_ISO27002", "Nombre_Control", "Descripcion", "Efectividad(0-100)"],

    "RIESGO_RESIDUAL": ["ID_Evaluacion", "ID_Riesgo", "ID_Activo", "Fecha", "Riesgo_Inherente", "Efectividad", "Riesgo_Residual"],

    "HISTORICO": ["ID_Evaluacion", "Fecha", "Total_Activos", "Riesgo_Promedio_Residual", "Riesgos_Altos", "Riesgos_Medios", "Riesgos_Bajos"],
    "DASHBOARD": ["KPI", "Valor"],
    "COMPARATIVO": ["Evaluacion_A", "Evaluacion_B", "Metrica", "Valor_A", "Valor_B", "Diferencia"],
}

def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

def main():
    if os.path.exists(EXCEL_PATH):
        print(f"Ya existe {EXCEL_PATH}. No lo sobrescribo.")
        return

    wb = Workbook()
    wb.remove(wb.active)

    for name, headers in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        autosize(ws)

    wb.save(EXCEL_PATH)
    print(f"Creado {EXCEL_PATH} con pestañas completas.")

if __name__ == "__main__":
    main()

